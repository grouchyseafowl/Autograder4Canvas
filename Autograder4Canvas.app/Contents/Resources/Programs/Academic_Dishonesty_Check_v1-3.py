import os
import platform
import requests
import re
from typing import List, Dict, Any, Tuple
from collections import defaultdict
import csv
from datetime import datetime
from pathlib import Path

# ======================
# CONFIGURATION
# ======================
CANVAS_BASE_URL = "https://cabrillo.instructure.com"
API_TOKEN = os.getenv("CANVAS_API_TOKEN")

if not API_TOKEN:
    raise ValueError("❌ Missing CANVAS_API_TOKEN environment variable")

HEADERS = {
    "Authorization": f"Bearer {API_TOKEN}",
    "Content-Type": "application/json"
}

def get_output_base_dir() -> Path:
    """Get base output directory in a cross-platform way."""
    system = platform.system()
    
    if os.path.isdir("/output"):
        return Path("/output")
    
    if system == "Windows":
        documents = Path(os.environ.get("USERPROFILE", Path.home())) / "Documents"
    else:
        documents = Path.home() / "Documents"
    
    return documents / "Grading_Rationales"

# Thresholds for flagging
MIN_WORD_COUNT = 50
MIN_FILE_SIZE = 1024  # 1KB in bytes
DUPLICATE_SIMILARITY_THRESHOLD = 0.85

# AI detection patterns - clichéd transitions and meta-commentary
AI_TRANSITIONS = [
    "it is important to note", "it should be noted", "it is worth noting",
    "in conclusion", "to sum up", "in summary", "to summarize",
    "furthermore", "moreover", "additionally", "in addition",
    "this essay will explore", "this paper will examine", "this section will discuss",
    "as previously mentioned", "as stated above", "as discussed earlier",
    "on the one hand", "on the other hand", "conversely", "however",
    "firstly", "secondly", "thirdly", "lastly", "finally"
]

# Hedge phrases and formal markers
HEDGE_PHRASES = [
    "it can be argued that", "it could be said that", "one might argue",
    "it is possible that", "arguably", "presumably", "ostensibly",
    "it appears that", "it seems that", "tends to suggest"
]

# Vocabulary inflation - unnecessarily complex words
INFLATED_VOCAB = [
    ("utilize", "use"), ("demonstrate", "show"), ("individuals", "people"),
    ("commence", "start"), ("terminate", "end"), ("endeavor", "try"),
    ("facilitate", "help"), ("implement", "do"), ("ascertain", "find out"),
    ("optimal", "best"), ("subsequent", "next"), ("prior to", "before"),
    ("in order to", "to"), ("due to the fact that", "because")
]

# Generic/vague phrases lacking specificity
GENERIC_PHRASES = [
    "many things", "various aspects", "in today's society", "in today's world",
    "throughout history", "since the beginning of time", "it can be said",
    "some people believe", "it is believed that", "one could argue",
    "studies show", "research indicates", "experts say", "it has been shown",
    "this shows that", "this proves that", "overall", "basically",
    "a variety of", "a number of", "plays a crucial role", "of paramount importance"
]

# Balanced/false equivalence markers
BALANCE_MARKERS = [
    "both sides", "both perspectives", "different viewpoints", "various opinions",
    "while some argue", "others contend", "there are arguments for and against"
]

# Passive voice patterns
PASSIVE_PATTERNS = [
    r"\b(is|are|was|were|be|been|being)\s+\w+ed\b",
    r"\b(can|could|may|might|should|would)\s+be\s+\w+ed\b"
]

# Personal/embodied markers (lack of these is a flag)
PERSONAL_MARKERS = [
    r"\bi\s", r"\bmy\b", r"\bme\b", r"\bmine\b",
    r"\bwe\b", r"\bour\b", r"\bus\b",
    r"as a\b", r"being a\b", r"growing up"
]

# Emotional/vulnerable language
EMOTIONAL_MARKERS = [
    "i feel", "i felt", "i was", "i struggled", "i realized", "i noticed",
    "confused", "frustrated", "excited", "surprised", "shocked", "angry",
    "it hurt", "i cried", "i laughed", "nervous", "scared", "proud"
]

# Generic filename patterns
GENERIC_FILENAMES = [
    r"^untitled\d*\..*$", r"^document\d*\..*$", r"^assignment\d*\..*$",
    r"^file\d*\..*$", r"^download\d*\..*$", r"^new file\d*\..*$",
    r"^copy of .*$"
]

def count_words(text: str) -> int:
    """Count words in text."""
    if not text:
        return 0
    return len(text.split())

def calculate_text_similarity(text1: str, text2: str) -> float:
    """Calculate simple similarity ratio between two texts."""
    if not text1 or not text2:
        return 0.0
    
    t1 = text1.lower().strip()
    t2 = text2.lower().strip()
    
    if t1 == t2:
        return 1.0
    
    words1 = set(t1.split())
    words2 = set(t2.split())
    
    if not words1 or not words2:
        return 0.0
    
    intersection = words1.intersection(words2)
    union = words1.union(words2)
    
    return len(intersection) / len(union) if union else 0.0

def get_all_discussion_topics(course_id: int) -> List[Dict]:
    """Fetch all discussion topics in the course."""
    print("💬 Fetching all discussion topics...")
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/discussion_topics"
    params = {"per_page": 100}
    response = requests.get(url, headers=HEADERS, params=params)

    if response.status_code != 200:
        print(f"❌ Failed to fetch discussion topics: {response.text}")
        return []

    data = response.json()
    return data if isinstance(data, list) else []


def fetch_discussion_entries(course_id: int, topic_id: int) -> List[Dict]:
    """Fetch all entries (including replies) for a discussion topic."""
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/discussion_topics/{topic_id}/view"
    resp = requests.get(url, headers=HEADERS)
    if resp.status_code != 200:
        print(f"⚠️  Failed to fetch discussion {topic_id} (HTTP {resp.status_code})")
        if "text/html" in resp.headers.get("content-type", ""):
            print("   💡 Likely cause: Canvas returned login page (check token/URL)")
        return []
    try:
        data = resp.json()
        entries = data.get("view", [])
    except Exception as e:
        print(f"   ❌ JSON decode error: {e}")
        return []
    return entries


def analyze_discussion_posts(course_id: int, topic: Dict, students: List[Dict]) -> Tuple[Dict[int, List[str]], Dict[int, str]]:
    """
    Analyze all posts in a discussion topic for academic dishonesty flags.
    Returns: (Dict mapping user_id to list of flags, Dict mapping user_id to post text)
    """
    topic_id = topic.get("id")
    topic_name = topic.get("title", f"Discussion {topic_id}")
    
    print(f"   💬 Analyzing discussion: {topic_name}")
    
    # Get all entries for this discussion
    entries = fetch_discussion_entries(course_id, topic_id)
    if not entries:
        return {}, {}
    
    # Flatten replies to get all posts
    all_posts = []
    def flatten_posts(posts):
        for post in posts:
            all_posts.append(post)
            if "replies" in post:
                flatten_posts(post["replies"])
    
    flatten_posts(entries)
    
    student_flags = {}
    post_texts = {}
    
    # Create a mapping of student user IDs for quick lookup
    student_user_ids = {student.get("user_id") for student in students if student.get("user_id")}
    
    for post in all_posts:
        user_id = post.get("user_id")
        
        # Only analyze posts from enrolled students
        if user_id not in student_user_ids:
            continue
        
        message = post.get("message", "")
        if message:
            # Create a submission-like object for compatibility with analyze_submission
            submission = {
                "body": message,
                "user_id": user_id,
                "workflow_state": "submitted"
            }
            
            flags = analyze_submission(submission, all_posts)
            if flags:
                if user_id not in student_flags:
                    student_flags[user_id] = []
                student_flags[user_id].extend(flags)
            
            # Store post text for cross-analysis
            if user_id not in post_texts:
                post_texts[user_id] = ""
            post_texts[user_id] += " " + message
    
    return student_flags, post_texts

def check_ai_transitions(text: str) -> int:
    """Count clichéd transitions and meta-commentary."""
    if not text:
        return 0
    
    text_lower = text.lower()
    count = 0
    
    for phrase in AI_TRANSITIONS:
        count += text_lower.count(phrase)
    
    return count

def check_hedge_phrases(text: str) -> int:
    """Count hedge phrases indicating avoidance of stance."""
    if not text:
        return 0
    
    text_lower = text.lower()
    count = 0
    
    for phrase in HEDGE_PHRASES:
        count += text_lower.count(phrase)
    
    return count

def check_inflated_vocabulary(text: str) -> List[str]:
    """Detect unnecessarily complex vocabulary."""
    if not text:
        return []
    
    text_lower = text.lower()
    found = []
    
    for complex_word, simple_word in INFLATED_VOCAB:
        if complex_word in text_lower:
            found.append(f"{complex_word} (vs {simple_word})")
    
    return found

def check_generic_content(text: str) -> int:
    """Count generic/vague phrases lacking specificity."""
    if not text:
        return 0
    
    text_lower = text.lower()
    count = 0
    
    for phrase in GENERIC_PHRASES:
        count += text_lower.count(phrase)
    
    return count

def check_balance_markers(text: str) -> int:
    """Check for false equivalence/over-balanced writing."""
    if not text:
        return 0
    
    text_lower = text.lower()
    count = 0
    
    for marker in BALANCE_MARKERS:
        count += text_lower.count(marker)
    
    return count

def check_passive_voice(text: str) -> int:
    """Count passive voice constructions."""
    if not text:
        return 0
    
    count = 0
    for pattern in PASSIVE_PATTERNS:
        count += len(re.findall(pattern, text, re.IGNORECASE))
    
    return count

def check_personal_markers(text: str) -> int:
    """Count first-person and embodied language."""
    if not text:
        return 0
    
    count = 0
    text_lower = text.lower()
    
    for pattern in PERSONAL_MARKERS:
        count += len(re.findall(pattern, text_lower))
    
    return count

def check_emotional_markers(text: str) -> int:
    """Count emotional/vulnerable language."""
    if not text:
        return 0
    
    text_lower = text.lower()
    count = 0
    
    for marker in EMOTIONAL_MARKERS:
        count += text_lower.count(marker)
    
    return count

def check_paragraph_uniformity(text: str) -> float:
    """Check if paragraphs are suspiciously uniform in length."""
    if not text:
        return 0.0
    
    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
    
    if len(paragraphs) < 3:
        return 0.0
    
    lengths = [len(p.split()) for p in paragraphs]
    avg_length = sum(lengths) / len(lengths)
    
    # Calculate coefficient of variation
    variance = sum((l - avg_length) ** 2 for l in lengths) / len(lengths)
    std_dev = variance ** 0.5
    
    if avg_length == 0:
        return 0.0
    
    coefficient_of_variation = std_dev / avg_length
    
    # Low variation = high uniformity (suspicious)
    return 1 - coefficient_of_variation if coefficient_of_variation < 1 else 0

def check_repetitive_reasoning(text: str) -> bool:
    """Detect repetitive or circular reasoning."""
    if not text:
        return False
    
    sentences = re.split(r'[.!?]+', text)
    sentences = [s.strip().lower() for s in sentences if len(s.strip()) > 20]
    
    if len(sentences) < 5:
        return False
    
    # Check for high similarity between sentences
    similar_pairs = 0
    for i in range(len(sentences) - 1):
        for j in range(i + 1, len(sentences)):
            similarity = calculate_text_similarity(sentences[i], sentences[j])
            if similarity > 0.7:
                similar_pairs += 1
    
    # If more than 20% of sentence pairs are highly similar
    total_pairs = (len(sentences) * (len(sentences) - 1)) / 2
    return (similar_pairs / total_pairs) > 0.2 if total_pairs > 0 else False

def check_copy_paste_indicators(text: str) -> List[str]:
    """Check for copy-paste indicators."""
    if not text:
        return []
    
    indicators = []
    
    if re.search(r'  +', text):
        indicators.append("Multiple consecutive spaces")
    
    if re.search(r'\n{3,}', text):
        indicators.append("Excessive line breaks")
    
    if '\t' in text and '    ' in text:
        indicators.append("Mixed tab/space formatting")
    
    if re.search(r'\x00|\x0c|\ufeff', text):
        indicators.append("Special formatting characters")
    
    return indicators

def check_sentence_completeness(text: str) -> float:
    """Calculate ratio of complete sentences."""
    if not text:
        return 0.0
    
    sentences = re.split(r'[.!?]+', text)
    sentences = [s.strip() for s in sentences if s.strip()]
    
    if not sentences:
        return 0.0
    
    complete_count = 0
    for sentence in sentences:
        words = sentence.split()
        if len(words) >= 8 and sentence[0].isupper():
            complete_count += 1
    
    return complete_count / len(sentences)

def check_generic_filename(filename: str) -> bool:
    """Check if filename matches generic patterns."""
    if not filename:
        return False
    
    filename_lower = filename.lower()
    
    for pattern in GENERIC_FILENAMES:
        if re.match(pattern, filename_lower):
            return True
    
    return False

def analyze_submission(submission: Dict[str, Any], all_submissions: List[Dict[str, Any]]) -> List[str]:
    """
    Analyze submission and return list of flags.
    Returns: list_of_flags
    """
    flags = []
    
    # Check text body
    body = submission.get("body", "")
    if body:
        word_count = count_words(body)
        
        # 1. Length check
        if word_count > 0 and word_count < MIN_WORD_COUNT:
            flags.append(f"Very short text ({word_count} words)")
        
        # 2. Overly polished: clichéd transitions
        transition_count = check_ai_transitions(body)
        if transition_count >= 3:
            flags.append(f"Clichéd transitions ({transition_count} instances)")
        
        # 3. Hedge phrases (avoiding stance)
        hedge_count = check_hedge_phrases(body)
        if hedge_count >= 2:
            flags.append(f"Excessive hedging ({hedge_count} hedge phrases)")
        
        # 4. Vocabulary inflation
        inflated = check_inflated_vocabulary(body)
        if len(inflated) >= 3:
            flags.append(f"Inflated vocabulary: {', '.join(inflated[:3])}")
        
        # 5. Generic/vague content
        generic_count = check_generic_content(body)
        if generic_count >= 3:
            flags.append(f"Generic/vague content ({generic_count} vague phrases)")
        
        # 6. Over-balanced writing
        balance_count = check_balance_markers(body)
        if balance_count >= 2:
            flags.append(f"Over-balanced/false equivalence ({balance_count} markers)")
        
        # 7. Excessive passive voice
        passive_count = check_passive_voice(body)
        total_sentences = len(re.split(r'[.!?]+', body))
        if total_sentences > 5 and (passive_count / total_sentences) > 0.4:
            flags.append(f"Excessive passive voice ({int((passive_count/total_sentences)*100)}%)")
        
        # 8. Lack of personal markers
        personal_count = check_personal_markers(body)
        if word_count > 200 and personal_count < 3:
            flags.append("Lacks personal/embodied language")
        
        # 9. Lack of emotional markers
        emotional_count = check_emotional_markers(body)
        if word_count > 200 and emotional_count == 0:
            flags.append("No emotional/vulnerable language")
        
        # 10. Uniform paragraph structure
        uniformity = check_paragraph_uniformity(body)
        if uniformity > 0.8:
            flags.append("Suspiciously uniform paragraphs (mechanical structure)")
        
        # 11. Repetitive reasoning
        if check_repetitive_reasoning(body):
            flags.append("Repetitive/circular reasoning detected")
        
        # 12. Copy-paste indicators
        copy_indicators = check_copy_paste_indicators(body)
        if copy_indicators:
            flags.append(f"Copy-paste indicators: {', '.join(copy_indicators)}")
        
        # 13. Suspiciously polished (complete sentences for notes)
        completeness = check_sentence_completeness(body)
        if completeness > 0.8 and word_count > 100:
            flags.append(f"Suspiciously polished ({int(completeness*100)}% complete sentences)")
        
        # 14. Check for duplicates
        current_user_id = submission.get("user_id")
        for other_sub in all_submissions:
            other_user_id = other_sub.get("user_id")
            if other_user_id == current_user_id:
                continue
            
            other_body = other_sub.get("body", "")
            if other_body:
                similarity = calculate_text_similarity(body, other_body)
                if similarity >= DUPLICATE_SIMILARITY_THRESHOLD:
                    other_user_name = other_sub.get("user", {}).get("name", f"User {other_user_id}")
                    flags.append(f"Highly similar to {other_user_name}'s submission ({int(similarity*100)}% match)")
                    break
    
    # Check attachments
    attachments = submission.get("attachments", [])
    if attachments:
        for file in attachments:
            file_size = file.get("size", 0)
            file_name = file.get("filename", "unknown")
            
            if file_size > 0 and file_size < MIN_FILE_SIZE:
                flags.append(f"Small file '{file_name}' ({file_size} bytes)")
            
            if check_generic_filename(file_name):
                flags.append(f"Generic filename: '{file_name}'")
    
    url = submission.get("url")
    if url and not body and not attachments:
        flags.append("URL with no description")
    
    return flags

def get_active_students(course_id: int) -> List[Dict]:
    """Fetch active student enrollments."""
    print("📥 Fetching active student enrollments...")
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/enrollments"
    params = {"type": ["StudentEnrollment"], "state": ["active"], "per_page": 100}
    response = requests.get(url, headers=HEADERS, params=params)

    if not response.headers.get('Content-Type', '').startswith('application/json'):
        print("⚠️ Received non-JSON response")
        return []

    if response.status_code != 200:
        print(f"❌ Failed to fetch enrollments: {response.text}")
        return []

    data = response.json()
    return data if isinstance(data, list) else []

def get_all_assignments(course_id: int) -> List[Dict]:
    """Fetch all assignments in the course."""
    print("📚 Fetching all assignments...")
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/assignments"
    params = {"per_page": 100}
    response = requests.get(url, headers=HEADERS, params=params)

    if response.status_code != 200:
        print(f"❌ Failed to fetch assignments: {response.text}")
        return []

    data = response.json()
    return data if isinstance(data, list) else []

def get_submissions(course_id: int, assignment_id: int) -> Dict[int, Dict]:
    """Fetch submissions for an assignment."""
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions"
    params = {"include": ["attachments", "user"], "per_page": 100}
    response = requests.get(url, headers=HEADERS, params=params)

    if response.status_code != 200:
        return {}

    data = response.json()
    if not isinstance(data, list):
        return {}

    return {sub.get("user_id"): sub for sub in data if sub.get("user_id")}

def analyze_assignment(course_id: int, assignment: Dict, students: List[Dict]) -> Tuple[Dict[int, List[str]], Dict[int, str]]:
    """
    Analyze all submissions for an assignment.
    Returns: (Dict mapping user_id to list of flags, Dict mapping user_id to submission body)
    """
    assignment_id = assignment.get("id")
    assignment_name = assignment.get("name", f"Assignment {assignment_id}")
    
    print(f"   📝 Analyzing: {assignment_name}")
    
    submissions = get_submissions(course_id, assignment_id)
    if not submissions:
        return {}, {}
    
    all_submissions_list = list(submissions.values())
    student_flags = {}
    submission_texts = {}
    
    for student in students:
        user_id = student.get("user_id")
        if not user_id:
            continue
        
        submission = submissions.get(user_id)
        if submission and submission.get("workflow_state") != "unsubmitted":
            flags = analyze_submission(submission, all_submissions_list)
            if flags:
                student_flags[user_id] = flags
            
            # Store submission text for cross-assignment analysis
            body = submission.get("body", "")
            if body:
                submission_texts[user_id] = body
    
    return student_flags, submission_texts

def check_cross_assignment_similarity(all_data: Dict, student_names: Dict[int, str]) -> Tuple[Dict, Dict]:
    """
    Check for suspicious similarities across assignments.
    Returns: (student_self_plagiarism_dict, cross_student_plagiarism_dict)
    """
    print("\n🔍 Checking for cross-assignment similarities...")
    
    # Build structure: {user_id: {assignment_id: text}}
    student_submissions = defaultdict(dict)
    
    for assignment_id, assignment_data in all_data.items():
        for user_id, text in assignment_data.get("submission_texts", {}).items():
            if text and len(text.split()) > 50:  # Only check substantial submissions
                student_submissions[user_id][assignment_id] = text
    
    # Check 1: Self-plagiarism (student reusing their own work)
    self_plagiarism = {}
    
    for user_id, assignments in student_submissions.items():
        assignment_ids = list(assignments.keys())
        similar_pairs = []
        
        for i in range(len(assignment_ids)):
            for j in range(i + 1, len(assignment_ids)):
                aid1, aid2 = assignment_ids[i], assignment_ids[j]
                text1, text2 = assignments[aid1], assignments[aid2]
                
                similarity = calculate_text_similarity(text1, text2)
                if similarity >= 0.7:  # 70% threshold for self-plagiarism
                    assignment_name1 = all_data[aid1]["name"]
                    assignment_name2 = all_data[aid2]["name"]
                    similar_pairs.append({
                        "assignment1": assignment_name1,
                        "assignment1_id": aid1,
                        "assignment2": assignment_name2,
                        "assignment2_id": aid2,
                        "similarity": similarity
                    })
        
        if similar_pairs:
            self_plagiarism[user_id] = similar_pairs
    
    # Check 2: Cross-student plagiarism (different students, different assignments)
    cross_student_plagiarism = []
    
    # Build list of all (user_id, assignment_id, text) tuples
    all_submissions = []
    for user_id, assignments in student_submissions.items():
        for assignment_id, text in assignments.items():
            all_submissions.append((user_id, assignment_id, text))
    
    # Compare across students and assignments
    checked_pairs = set()
    
    for i in range(len(all_submissions)):
        user1, aid1, text1 = all_submissions[i]
        
        for j in range(i + 1, len(all_submissions)):
            user2, aid2, text2 = all_submissions[j]
            
            # Skip if same student or same assignment
            if user1 == user2 or aid1 == aid2:
                continue
            
            # Skip if we've already checked this pair
            pair_key = tuple(sorted([(user1, aid1), (user2, aid2)]))
            if pair_key in checked_pairs:
                continue
            checked_pairs.add(pair_key)
            
            similarity = calculate_text_similarity(text1, text2)
            if similarity >= 0.75:  # 75% threshold for cross-student
                cross_student_plagiarism.append({
                    "student1": student_names.get(user1, f"User {user1}"),
                    "student1_id": user1,
                    "assignment1": all_data[aid1]["name"],
                    "assignment1_id": aid1,
                    "student2": student_names.get(user2, f"User {user2}"),
                    "student2_id": user2,
                    "assignment2": all_data[aid2]["name"],
                    "assignment2_id": aid2,
                    "similarity": similarity
                })
    
    print(f"   Found {len(self_plagiarism)} students with self-plagiarism")
    print(f"   Found {len(cross_student_plagiarism)} cross-student similarities")
    
    return self_plagiarism, cross_student_plagiarism

def generate_report(course_id: int, course_name: str, all_data: Dict, self_plagiarism: Dict, cross_student_plagiarism: List, discussion_data: Dict = None):
    """Generate comprehensive report including assignments and discussion forums."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Cross-platform output base
    BASE_DIR = get_output_base_dir()

    # Academic Dishonesty outputs with CSV/Excel subfolders
    output_dir = BASE_DIR / "Academic Dishonesty Reports"
    csv_dir = output_dir / "csv"
    excel_dir = output_dir / "excel"
    csv_dir.mkdir(parents=True, exist_ok=True)
    excel_dir.mkdir(parents=True, exist_ok=True)
    
    # Console report
    print("\n" + "="*80)
    print(f"📊 FLAG REPORT: {course_name}")
    print("="*80)
    
    # Student-centric view for assignments
    print("\n" + "="*80)
    print("👥 FLAGS BY STUDENT (ASSIGNMENTS)")
    print("="*80)
    
    student_summary = defaultdict(lambda: {"total_flags": 0, "assignments": {}})
    
    for assignment_id, assignment_data in all_data.items():
        assignment_name = assignment_data["name"]
        for user_id, flags in assignment_data["flags"].items():
            student_name = assignment_data["student_names"].get(user_id, f"User {user_id}")
            student_summary[user_id]["name"] = student_name
            student_summary[user_id]["total_flags"] += len(flags)
            student_summary[user_id]["assignments"][assignment_name] = flags
    
    # Sort by total flags
    sorted_students = sorted(student_summary.items(), key=lambda x: x[1]["total_flags"], reverse=True)
    
    for user_id, data in sorted_students:
        print(f"\n👤 {data['name']} (ID: {user_id})")
        print(f"   Total flags: {data['total_flags']} across {len(data['assignments'])} assignments")
        print("-" * 80)
        for assignment_name, flags in data["assignments"].items():
            print(f"\n   📋 {assignment_name}:")
            for flag in flags:
                print(f"      • {flag}")
    
    # Add section for discussion forums if available
    if discussion_data:
        print("\n" + "="*80)
        print("💬 FLAGS IN DISCUSSION FORUMS")
        print("="*80)
        
        discussion_summary = defaultdict(lambda: {"total_flags": 0, "discussions": {}})
        
        for topic_id, topic_data in discussion_data.items():
            topic_name = topic_data["name"]
            for user_id, flags in topic_data["flags"].items():
                student_name = topic_data["student_names"].get(user_id, f"User {user_id}")
                discussion_summary[user_id]["name"] = student_name
                discussion_summary[user_id]["total_flags"] += len(flags)
                discussion_summary[user_id]["discussions"][topic_name] = flags
        
        # Sort by total flags
        sorted_discussion_students = sorted(discussion_summary.items(), key=lambda x: x[1]["total_flags"], reverse=True)
        
        for user_id, data in sorted_discussion_students:
            print(f"\n👤 {data['name']} (ID: {user_id})")
            print(f"   Total discussion flags: {data['total_flags']} across {len(data['discussions'])} discussions")
            print("-" * 80)
            for topic_name, flags in data["discussions"].items():
                print(f"\n   💬 {topic_name}:")
                for flag in flags:
                    print(f"      • {flag}")

    # CSV Export (detailed list for assignments)
    csv_path = csv_dir / f"canvas_flag_report_{course_id}_{timestamp}.csv"
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Type", "Student Name", "User ID", "Item", "Item ID", "Flag"])
        
        # Write assignment flags
        for assignment_id, assignment_data in all_data.items():
            assignment_name = assignment_data["name"]
            for user_id, flags in assignment_data["flags"].items():
                student_name = assignment_data["student_names"].get(user_id, f"User {user_id}")
                for flag in flags:
                    writer.writerow(["Assignment", student_name, user_id, assignment_name, assignment_id, flag])
        
        # Write discussion flags
        if discussion_data:
            for topic_id, topic_data in discussion_data.items():
                topic_name = topic_data["name"]
                for user_id, flags in topic_data["flags"].items():
                    student_name = topic_data["student_names"].get(user_id, f"User {user_id}")
                    for flag in flags:
                        writer.writerow(["Discussion", student_name, user_id, topic_name, topic_id, flag])
    
    print(f"\n✅ CSV report saved: {csv_path.name}")
    
    # Excel Export with multiple sheets (simplified version)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        excel_path = excel_dir / f"canvas_flag_report_{course_id}_{timestamp}.xlsx"
        wb = openpyxl.Workbook()
        
        # Sheet 1: Combined Summary
        ws_summary = wb.active
        ws_summary.title = "Combined Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ["Type", "Student Name", "User ID", "Item", "Item ID", "Flag Count", "Flags"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 2
        # Add assignment data
        for assignment_id, assignment_data in all_data.items():
            for user_id, flags in assignment_data["flags"].items():
                student_name = assignment_data["student_names"].get(user_id, f"User {user_id}")
                ws_summary.cell(row=row, column=1, value="Assignment")
                ws_summary.cell(row=row, column=2, value=student_name)
                ws_summary.cell(row=row, column=3, value=user_id)
                ws_summary.cell(row=row, column=4, value=assignment_data["name"])
                ws_summary.cell(row=row, column=5, value=assignment_id)
                ws_summary.cell(row=row, column=6, value=len(flags))
                ws_summary.cell(row=row, column=7, value="; ".join(flags))
                row += 1
        
        # Add discussion data
        if discussion_data:
            for topic_id, topic_data in discussion_data.items():
                for user_id, flags in topic_data["flags"].items():
                    student_name = topic_data["student_names"].get(user_id, f"User {user_id}")
                    ws_summary.cell(row=row, column=1, value="Discussion")
                    ws_summary.cell(row=row, column=2, value=student_name)
                    ws_summary.cell(row=row, column=3, value=user_id)
                    ws_summary.cell(row=row, column=4, value=topic_data["name"])
                    ws_summary.cell(row=row, column=5, value=topic_id)
                    ws_summary.cell(row=row, column=6, value=len(flags))
                    ws_summary.cell(row=row, column=7, value="; ".join(flags))
                    row += 1
        
        # Auto-adjust column widths
        for col in range(1, 8):
            ws_summary.column_dimensions[get_column_letter(col)].width = 20
        
        wb.save(excel_path)
        print(f"✅ Excel report saved: {excel_path.name}")
        
    except ImportError:
        print("⚠️ openpyxl not installed. Excel export skipped.")
        print("   Install with: pip install openpyxl")
    

    
    # Summary statistics
    print("\n" + "="*80)
    print("📈 SUMMARY STATISTICS")
    print("="*80)
    
    total_assignments = len(all_data)
    total_flagged_assignments = sum(1 for a in all_data.values() if a["flags"])
    total_flags = sum(len(a["flags"]) for a in all_data.values())
    total_students_flagged = len(student_summary)
    
    print(f"Total assignments analyzed: {total_assignments}")
    print(f"Assignments with flags: {total_flagged_assignments}")
    print(f"Total flags issued: {total_flags}")
    print(f"Students with flags: {total_students_flagged}")
    
    if total_students_flagged > 0:
        avg_flags_per_student = sum(s["total_flags"] for s in student_summary.values()) / total_students_flagged
        print(f"Average flags per flagged student: {avg_flags_per_student:.1f}")

    # Final output location reminder
    print(f"\n📁 Reports were exported to:")
    print(f"   CSV: {csv_path}")
    if 'excel_path' in locals():
        print(f"   Excel: {excel_path}")

def main():
    print("🎓 Academic Dishonesty Flag Generator (for Canvas)")
    print("Designed for detecting AI-generated content, plagiarism, and bad faith work.")
    print("NO GRADES WILL BE SUBMITTED.\n")

    try:
        course_id = int(input("Enter Course ID: ").strip())
    except ValueError:
        print("❌ Invalid course ID.")
        return

    # Verify API access
    print("🔍 Verifying Canvas API access...")
    test_url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}"
    test_resp = requests.get(test_url, headers=HEADERS)
    
    if not test_resp.headers.get('Content-Type', '').startswith('application/json'):
        print("❌ CRITICAL: Received HTML response.")
        print("   Check your token, URL, and internet connection.")
        return
    
    if test_resp.status_code != 200:
        print(f"❌ Could not access course: {test_resp.text}")
        return
    
    course_data = test_resp.json()
    course_name = course_data.get("name", f"Course {course_id}")
    print(f"✅ Connected to: {course_name}\n")

    # Get students
    students = get_active_students(course_id)
    if not students:
        print("🛑 No active students found. Exiting.")
        return
    
    print(f"✅ Found {len(students)} active students\n")

    # Ask if user wants to analyze all assignments or specific ones
    print("\n📋 Assignment Selection Options:")
    print("   [1] Analyze ALL assignments in this course")
    print("   [2] Analyze specific assignments by ID")
    print("   [3] Analyze assignments filtered by keyword")
    choice = input("\nChoose an option (1/2/3, default=1): ").strip() or "1"

    assignments = get_all_assignments(course_id)
    if not assignments:
        print("🛑 No assignments found. Exiting.")
        return

    print(f"\n✅ Found {len(assignments)} total assignments in the course")

    if choice == "1":
        # Use all assignments
        print("\n✅ Selected: ALL assignments will be analyzed")
        
    elif choice == "2":
        # Manual assignment ID entry
        assignment_input = input("\nEnter specific Assignment ID(s) (space-separated): ").strip()
        try:
            assignment_ids = [int(x.strip()) for x in assignment_input.split() if x.strip()]
            if not assignment_ids:
                print("❌ No assignment IDs provided. Using all assignments instead.")
            else:
                # Filter assignments list to only include specified IDs
                filtered_assignments = [a for a in assignments if a.get("id") in assignment_ids]
                if len(filtered_assignments) != len(assignment_ids):
                    print(f"⚠️  Some assignment IDs not found. Analyzing {len(filtered_assignments)} valid assignments.")
                assignments = filtered_assignments
        except ValueError:
            print("❌ Invalid assignment ID format. Using all assignments instead.")
            
    elif choice == "3":
        # Filter by keyword
        filter_keyword = input("\nEnter keyword to filter assignments (e.g., 'essay', 'reflection'): ").strip().lower()
        if filter_keyword:
            assignments = [a for a in assignments if filter_keyword in a.get("name", "").lower()]
            print(f"📌 Filtered to {len(assignments)} assignments containing '{filter_keyword}'")
        else:
            print("⚠️  No keyword provided. Using all assignments.")

    # Ask if user wants to analyze discussion forums
    if choice == "1":
        print("\n💬 Fetching ALL discussion topics (auto-selected with 'Analyze ALL')...")
        analyze_discussions = True
        discussion_topics = get_all_discussion_topics(course_id)
        print(f"✅ Found {len(discussion_topics)} discussion topics\n")
    else:
        # Only ask for confirmation and filtering in manual modes ([2] or [3])
        analyze_discussions = input("Also analyze discussion forums for academic dishonesty? (y/n, default=n): ").strip().lower() == 'y'
        if analyze_discussions:
            print("\n💬 Fetching discussion topics...")
            discussion_topics = get_all_discussion_topics(course_id)
            print(f"✅ Found {len(discussion_topics)} discussion topics\n")
            if discussion_topics:
                filter_discussions = input("Filter discussion topics? (y/n, default=n): ").strip().lower() == 'y'
                if filter_discussions:
                    filter_keyword = input("Enter keyword to filter discussions: ").strip().lower()
                    if filter_keyword:
                        discussion_topics = [t for t in discussion_topics if filter_keyword in t.get("title", "").lower()]
                        print(f"📌 Filtered to {len(discussion_topics)} discussions containing '{filter_keyword}'\n")
        else:
            discussion_topics = []

    # Analyze assignments
    print(f"\n🔍 Analyzing assignments ({len(assignments)} total)...")
    all_data = {}
    for idx, assignment in enumerate(assignments, 1):
        assignment_id = assignment.get("id")
        assignment_name = assignment.get("name", f"Assignment {assignment_id}")
        
        print(f"\n📝 [{idx}/{len(assignments)}] Analyzing: {assignment_name}")
        
        student_flags, submission_texts = analyze_assignment(course_id, assignment, students)
        
        # Get student names from submissions
        submissions = get_submissions(course_id, assignment_id)
        student_names = {}
        for user_id, sub in submissions.items():
            user_info = sub.get("user", {})
            student_names[user_id] = user_info.get("name", f"User {user_id}")
        
        all_data[assignment_id] = {
            "name": assignment_name,
            "flags": student_flags,
            "student_names": student_names,
            "submission_texts": submission_texts
        }
    
    print("\n✅ Assignment analysis complete!")
    
    # Analyze discussion topics if requested
    discussion_data = {}
    if analyze_discussions and discussion_topics:
        print("\n🔍 Analyzing discussion topics...")
        
        for topic in discussion_topics:
            topic_id = topic.get("id")
            topic_name = topic.get("title", f"Discussion {topic_id}")
            
            student_flags, post_texts = analyze_discussion_posts(course_id, topic, students)
            
            # Get student names from discussion posts
            student_names = {}
            entries = fetch_discussion_entries(course_id, topic_id)
            for entry in entries:
                user_id = entry.get("user_id")
                if user_id:
                    user_info = entry.get("user", {})
                    student_names[user_id] = user_info.get("name", f"User {user_id}")
            
            discussion_data[topic_id] = {
                "name": topic_name,
                "flags": student_flags,
                "student_names": student_names,
                "post_texts": post_texts
            }
        
        print("\n✅ Discussion forum analysis complete!")

    # Check for cross-assignment similarities
    student_names_global = {}
    for assignment_data in all_data.values():
        student_names_global.update(assignment_data.get("student_names", {}))
    
    # Also include discussion student names
    for topic_data in discussion_data.values():
        student_names_global.update(topic_data.get("student_names", {}))
    
    self_plagiarism, cross_student_plagiarism = check_cross_assignment_similarity(all_data, student_names_global)
    
    # Generate report
    generate_report(course_id, course_name, all_data, self_plagiarism, cross_student_plagiarism, discussion_data)
    
    print("\n" + "="*80)
    print("🏁 Report generation complete!")
    print("="*80)

    # ✅ Final output location instructions (cross-platform)
    # Define output_dir here since generate_report() has its own local scope
    output_dir = get_output_base_dir() / "Academic Dishonesty Reports"
    print(f"\n📁 All reports were exported to: {output_dir}")
    
    system = platform.system()
    if system == "Darwin":
        print(f"\n   🔍 To view in Finder:")
        print(f"       Open Finder → Go → Go to Folder… → paste the path above")
        print(f"\n   💻 To open in Terminal:")
        print(f'       open "{output_dir}"')
    elif system == "Windows":
        print(f"\n   🔍 To view in Explorer:")
        print(f"       Press Win+E → paste the path in the address bar")
        print(f"\n   💻 To open in Command Prompt:")
        print(f'       explorer "{output_dir}"')
    else:
        print(f"\n   🔍 To view in file manager:")
        print(f"       Open your file manager and navigate to the path above")
        print(f"\n   💻 To open in Terminal:")
        print(f'       xdg-open "{output_dir}"')

if __name__ == "__main__":
    main()