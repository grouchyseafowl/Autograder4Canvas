import os
import platform
import requests
import re
from typing import List, Dict, Any, Tuple
from collections import defaultdict
import csv
from datetime import datetime
from pathlib import Path

# ======================
# CONFIGURATION
# ======================
CANVAS_BASE_URL = "https://cabrillo.instructure.com"
API_TOKEN = os.getenv("CANVAS_API_TOKEN")

if not API_TOKEN:
    raise ValueError("❌ Missing CANVAS_API_TOKEN environment variable")

HEADERS = {
    "Authorization": f"Bearer {API_TOKEN}",
    "Content-Type": "application/json"
}

def get_config_dir() -> Path:
    """Get the configuration directory for storing settings."""
    system = platform.system()
    
    if system == "Windows":
        base = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local"))
        config_dir = base / "CanvasAutograder"
    elif system == "Darwin":
        config_dir = Path.home() / "Library" / "Application Support" / "CanvasAutograder"
    else:
        config_dir = Path.home() / ".config" / "CanvasAutograder"
    
    return config_dir


def get_output_base_dir() -> Path:
    """Get base output directory in a cross-platform way."""
    system = platform.system()
    
    # Check for /output directory first (container/deployment environment)
    if os.path.isdir("/output"):
        return Path("/output")
    
    # Check for custom output directory in JSON config (matches autograder_utils.py)
    config_file = get_config_dir() / "settings.json"
    if config_file.exists():
        try:
            import json
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                if "output_directory" in config:
                    custom_dir = Path(config["output_directory"])
                    if custom_dir.exists() or custom_dir.parent.exists():
                        custom_dir.mkdir(parents=True, exist_ok=True)
                        return custom_dir
        except Exception:
            pass  # Fall through to default
    
    # Default location
    if system == "Windows":
        documents = Path(os.environ.get("USERPROFILE", Path.home())) / "Documents"
    else:
        documents = Path.home() / "Documents"
    
    return documents / "Autograder Rationales"

# Thresholds for flagging
MIN_WORD_COUNT = 50
MIN_FILE_SIZE = 1024  # 1KB in bytes
DUPLICATE_SIMILARITY_THRESHOLD = 0.85

# Global settings (set during runtime)
CURRENT_PROFILE = None
VERIFY_CITATIONS = False  # Whether to verify citations against CrossRef/OpenLibrary

# ======================
# ASSIGNMENT TYPE PROFILES
# ======================
# These profiles configure which checks run and their sensitivity based on assignment type.
# Students may not always follow assignment instructions perfectly (e.g., spellchecking a draft),
# so profiles include tolerance settings and contextual notes for instructors.

ASSIGNMENT_PROFILES = {
    "notes_brainstorm": {
        "name": "Notes / Brainstorming / Outlines",
        "description": "Informal notes, brainstorming sessions, outlines, or pre-writing activities",
        "expectations": {
            "polish_level": "very_low",
            "personal_voice": "optional",
            "structure": "informal",
            "errors_expected": True,
            "fragments_expected": True,  # Fragments are EXPECTED - complete sentences are suspicious
            "paragraph_structure": False  # Should NOT have organized paragraphs
        },
        "checks_enabled": {
            "ai_transitions": False,          # Notes might have transitional thinking
            "hedge_phrases": False,            # Not relevant for notes
            "inflated_vocabulary": True,       # Still suspicious in notes
            "generic_phrases": False,          # Brainstorming is often generic
            "balance_markers": False,          # Not relevant
            "passive_voice": False,            # Not relevant for notes
            "personal_markers": False,         # Not required
            "emotional_markers": False,        # Not required
            "complete_sentences": True,        # ENABLED - to FLAG when too polished
            "paragraph_structure": True,       # ENABLED - to FLAG organized paragraphs
            "copy_paste": True,                # Still check for copy-paste
            "cross_submission": True,          # Check for reuse
            "essay_organization": True,        # Check for suspicious organization
            "headings_structure": True         # Check for AI-style headings
        },
        "thresholds": {
            "min_word_count": 20,
            "ai_transition_count": 999,        # Effectively disabled
            "passive_voice_percent": 100,      # Effectively disabled
            "complete_sentence_percent": 50,   # Flag if MORE than 50% complete (notes should be fragments)
            "paragraph_count_suspicious": 3    # Flag if has 3+ organized paragraphs
        },
        "invert_checks": {
            "complete_sentences": True,        # INVERT: flag HIGH completeness, not low
            "paragraph_structure": True        # INVERT: flag PRESENCE of structure, not absence
        },
        "instructor_notes": [
            "Notes/outlines should have INCOMPLETE sentences and informal structure",
            "Flag appears if submission looks like polished prose instead of notes",
            "Students told to use fragments should not submit paragraph-style work",
            "Repeated polished 'notes' submissions may indicate AI use"
        ]
    },
    
    "rough_draft": {
        "name": "Rough Draft / First Draft",
        "description": "Early drafts where errors and rough structure are expected",
        "expectations": {
            "polish_level": "low",
            "personal_voice": "developing",
            "structure": "emerging",
            "errors_expected": True,
            "fragments_ok": True
        },
        "checks_enabled": {
            "ai_transitions": True,            # Reduced sensitivity
            "hedge_phrases": False,            # Common in drafts
            "inflated_vocabulary": True,       # Suspicious in drafts
            "generic_phrases": True,           # Reduced sensitivity
            "balance_markers": False,          # Not relevant for drafts
            "passive_voice": False,            # Common in drafts
            "personal_markers": False,         # Developing
            "emotional_markers": False,        # Developing
            "complete_sentences": False,       # Drafts have fragments
            "copy_paste": True,                # Important to check
            "cross_submission": True           # Important to check
        },
        "thresholds": {
            "min_word_count": 30,
            "ai_transition_count": 5,          # Higher tolerance
            "passive_voice_percent": 60,       # Higher tolerance
            "complete_sentence_percent": 90    # Higher tolerance
        },
        "flags_to_note_not_flag": [
            "Some polish is normal - students may ignore 'don't spellcheck' instructions",
            "Focus on suspicious PATTERNS rather than polish level"
        ],
        "instructor_notes": [
            "Students often spellcheck/polish drafts despite instructions not to",
            "High polish alone is not definitive - look for AI-specific patterns",
            "Compare against final draft for authentic revision patterns"
        ]
    },
    
    "personal_reflection": {
        "name": "Personal Reflection / Response Paper",
        "description": "Personal essays, reflections, response papers requiring authentic voice",
        "expectations": {
            "polish_level": "medium",
            "personal_voice": "required",
            "structure": "flexible",
            "errors_expected": False,
            "fragments_ok": True  # Stylistic fragments OK
        },
        "checks_enabled": {
            "ai_transitions": True,            # Important check
            "hedge_phrases": True,             # Suspicious in personal writing
            "inflated_vocabulary": True,       # Very suspicious
            "generic_phrases": True,           # Important check
            "balance_markers": True,           # Suspicious in personal writing
            "passive_voice": True,             # Check but lower weight
            "personal_markers": True,          # CRITICAL - should be present
            "emotional_markers": True,         # CRITICAL - should be present
            "complete_sentences": True,        # Moderate sensitivity
            "copy_paste": True,
            "cross_submission": True
        },
        "thresholds": {
            "min_word_count": 50,
            "ai_transition_count": 3,
            "passive_voice_percent": 40,
            "complete_sentence_percent": 85
        },
        "required_markers": ["personal_voice", "emotional_language"],
        "absence_is_suspicious": True,
        "instructor_notes": [
            "ABSENCE of personal voice/emotion is a strong signal in reflections",
            "AI struggles to produce authentic personal narrative",
            "Look for specific, concrete personal examples vs. generic statements",
            "Even good students use 'I' and share genuine experiences in reflections"
        ]
    },
    
    "discussion_post": {
        "name": "Discussion Forum Post",
        "description": "Online discussion posts, typically requiring engagement with peers/readings",
        "expectations": {
            "polish_level": "medium_low",
            "personal_voice": "expected",
            "structure": "conversational",
            "errors_expected": False,
            "fragments_ok": True
        },
        "checks_enabled": {
            "ai_transitions": True,
            "hedge_phrases": True,             # Less suspicious in discussions
            "inflated_vocabulary": True,
            "generic_phrases": True,           # Important check
            "balance_markers": True,
            "passive_voice": False,            # Not relevant
            "personal_markers": True,          # Should be present
            "emotional_markers": True,         # Often present
            "complete_sentences": False,       # Conversational tone
            "copy_paste": True,
            "cross_submission": True           # Check for reuse across posts
        },
        "thresholds": {
            "min_word_count": 30,
            "ai_transition_count": 2,
            "passive_voice_percent": 50,
            "complete_sentence_percent": 90
        },
        "special_checks": {
            "references_classmates": "expected",     # Should mention peers
            "references_readings": "expected",       # Should cite course material
            "asks_questions": "common"               # Engagement indicator
        },
        "instructor_notes": [
            "Check if post references specific classmates or readings",
            "Generic posts that could apply to any class are suspicious",
            "AI posts often lack genuine questions or curiosity"
        ]
    },
    
    "analytical_essay": {
        "name": "Analytical / Argumentative Essay",
        "description": "Formal essays with thesis, analysis, and argument (first person may be allowed)",
        "expectations": {
            "polish_level": "high",
            "personal_voice": "optional",      # Depends on instructor preference
            "structure": "formal",
            "errors_expected": False,
            "fragments_ok": False
        },
        "checks_enabled": {
            "ai_transitions": True,            # Important check
            "hedge_phrases": True,             # Important check
            "inflated_vocabulary": True,       # Important check
            "generic_phrases": True,           # Very important
            "balance_markers": True,           # Check for false balance
            "passive_voice": True,             # Check but context-dependent
            "personal_markers": False,         # Not required
            "emotional_markers": False,        # Not required
            "complete_sentences": True,        # Expected to be polished
            "copy_paste": True,
            "cross_submission": True
        },
        "thresholds": {
            "min_word_count": 50,
            "ai_transition_count": 3,
            "passive_voice_percent": 45,
            "complete_sentence_percent": 80    # High polish expected
        },
        "content_checks": {
            "has_clear_thesis": "expected",
            "has_specific_evidence": "expected",
            "makes_original_argument": "expected"
        },
        "instructor_notes": [
            "High polish is EXPECTED - don't flag for polish alone",
            "Focus on: originality of argument, specific evidence, AI-specific phrases",
            "Check if thesis is genuinely argued or just stated",
            "AI essays often lack original insight even when well-structured"
        ]
    },
    
    "research_paper": {
        "name": "Research Paper with Citations",
        "description": "Formal research papers requiring sources and citations",
        "expectations": {
            "polish_level": "very_high",
            "personal_voice": "minimal",
            "structure": "formal",
            "errors_expected": False,
            "fragments_ok": False
        },
        "checks_enabled": {
            "ai_transitions": True,
            "hedge_phrases": True,
            "inflated_vocabulary": True,
            "generic_phrases": True,
            "balance_markers": True,
            "passive_voice": True,             # Common in research writing
            "personal_markers": False,
            "emotional_markers": False,
            "complete_sentences": True,
            "copy_paste": True,
            "cross_submission": True
        },
        "thresholds": {
            "min_word_count": 100,
            "ai_transition_count": 4,          # Some transitions expected
            "passive_voice_percent": 55,       # Higher tolerance for research
            "complete_sentence_percent": 75    # Very high polish expected
        },
        "special_checks": {
            "citations_present": "required",
            "citations_verifiable": "important",  # AI often fabricates sources
            "quotes_present": "expected"
        },
        "instructor_notes": [
            "AI often fabricates citations - VERIFY sources exist",
            "Check that quotes actually appear in cited sources",
            "Very high polish is expected - focus on content authenticity",
            "Look for vague references to 'studies show' without specifics"
        ]
    },
    
    "creative_writing": {
        "name": "Creative Writing / Fiction",
        "description": "Short stories, poetry, creative nonfiction",
        "expectations": {
            "polish_level": "varies",
            "personal_voice": "required",      # Unique voice expected
            "structure": "creative",
            "errors_expected": False,
            "fragments_ok": True               # Stylistic choice
        },
        "checks_enabled": {
            "ai_transitions": False,           # Different for creative work
            "hedge_phrases": False,            # Not relevant
            "inflated_vocabulary": True,       # Check for purple prose
            "generic_phrases": True,           # Clichés are a problem
            "balance_markers": False,
            "passive_voice": False,            # Stylistic choice
            "personal_markers": False,         # Depends on POV
            "emotional_markers": True,         # Should show not tell
            "complete_sentences": False,       # Stylistic fragments OK
            "copy_paste": True,
            "cross_submission": True
        },
        "thresholds": {
            "min_word_count": 50,
            "ai_transition_count": 999,        # Not relevant
            "passive_voice_percent": 100,      # Not relevant
            "complete_sentence_percent": 100   # Not relevant
        },
        "special_checks": {
            "unique_voice": "critical",
            "specific_sensory_details": "expected",
            "original_metaphors": "expected",
            "predictable_plot": "suspicious"
        },
        "instructor_notes": [
            "AI creative writing often feels generic and predictable",
            "Look for unique voice, surprising details, authentic emotion",
            "Clichéd descriptions and predictable plots are red flags",
            "Check for sensory specificity vs. generic descriptions"
        ]
    },
    
    "standard": {
        "name": "Standard Analysis (All Checks)",
        "description": "Default mode - runs all checks with standard sensitivity",
        "expectations": {
            "polish_level": "medium",
            "personal_voice": "context_dependent",
            "structure": "varies",
            "errors_expected": False,
            "fragments_ok": False
        },
        "checks_enabled": {
            "ai_transitions": True,
            "hedge_phrases": True,
            "inflated_vocabulary": True,
            "generic_phrases": True,
            "balance_markers": True,
            "passive_voice": True,
            "personal_markers": True,
            "emotional_markers": True,
            "complete_sentences": True,
            "copy_paste": True,
            "cross_submission": True
        },
        "thresholds": {
            "min_word_count": 50,
            "ai_transition_count": 3,
            "passive_voice_percent": 45,
            "complete_sentence_percent": 80
        },
        "instructor_notes": [
            "Default mode - use more specific profiles when possible",
            "All checks enabled with moderate sensitivity",
            "Review flags in context of actual assignment requirements"
        ]
    }
}

# Current analysis profile (set during runtime)
CURRENT_PROFILE = None

def get_profile(profile_key: str) -> dict:
    """Get an assignment profile by key."""
    return ASSIGNMENT_PROFILES.get(profile_key, ASSIGNMENT_PROFILES["standard"])

def is_check_enabled(check_name: str) -> bool:
    """Check if a specific analysis check is enabled for the current profile."""
    if CURRENT_PROFILE is None:
        return True  # Default to enabled
    profile = get_profile(CURRENT_PROFILE)
    return profile.get("checks_enabled", {}).get(check_name, True)

def get_threshold(threshold_name: str) -> int:
    """Get a threshold value for the current profile."""
    if CURRENT_PROFILE is None:
        # Return default thresholds
        defaults = {
            "min_word_count": 50,
            "ai_transition_count": 3,
            "passive_voice_percent": 45,
            "complete_sentence_percent": 80
        }
        return defaults.get(threshold_name, 50)
    
    profile = get_profile(CURRENT_PROFILE)
    return profile.get("thresholds", {}).get(threshold_name, 50)

def print_profile_selection_menu():
    """Print the assignment type selection menu."""
    print("\n" + "="*70)
    print("📝 ASSIGNMENT TYPE SELECTION")
    print("="*70)
    print()
    print("Select the type of assignment you're analyzing.")
    print("This adjusts which checks run and their sensitivity.")
    print()
    print("   [1] 📓 Notes / Brainstorming / Outlines")
    print("       Very permissive - only checks for copy-paste/plagiarism")
    print()
    print("   [2] 📝 Rough Draft / First Draft")
    print("       Tolerant of errors and rough structure")
    print("       Note: Students often polish drafts despite instructions")
    print()
    print("   [3] 💭 Personal Reflection / Response Paper")
    print("       Expects personal voice and genuine emotion")
    print("       ABSENCE of personal language is suspicious")
    print()
    print("   [4] 💬 Discussion Forum Post")
    print("       Should reference classmates/readings specifically")
    print()
    print("   [5] 📊 Analytical / Argumentative Essay")
    print("       Expects polish; checks for AI patterns and weak arguments")
    print("       First person may or may not be appropriate")
    print()
    print("   [6] 📚 Research Paper with Citations")
    print("       High polish expected; verifies citation authenticity")
    print()
    print("   [7] ✨ Creative Writing / Fiction")
    print("       Checks for unique voice and originality")
    print()
    print("   [8] 🔍 Standard Analysis (All Checks)")
    print("       Default mode with all checks enabled")
    print()

def select_assignment_profile() -> str:
    """Interactive selection of assignment profile."""
    print_profile_selection_menu()
    
    profile_map = {
        "1": "notes_brainstorm",
        "2": "rough_draft", 
        "3": "personal_reflection",
        "4": "discussion_post",
        "5": "analytical_essay",
        "6": "research_paper",
        "7": "creative_writing",
        "8": "standard"
    }
    
    while True:
        choice = input("Select assignment type (1-8, default=8): ").strip() or "8"
        
        if choice in profile_map:
            profile_key = profile_map[choice]
            profile = get_profile(profile_key)
            
            print()
            print(f"✅ Selected: {profile['name']}")
            print(f"   {profile['description']}")
            
            # Show what's enabled/disabled
            checks = profile.get("checks_enabled", {})
            enabled = [k for k, v in checks.items() if v]
            disabled = [k for k, v in checks.items() if not v]
            
            if disabled:
                print(f"\n   ℹ️  Checks DISABLED for this type:")
                for check in disabled[:5]:  # Show first 5
                    print(f"      • {check.replace('_', ' ').title()}")
                if len(disabled) > 5:
                    print(f"      • ... and {len(disabled) - 5} more")
            
            # Show instructor notes
            notes = profile.get("instructor_notes", [])
            if notes:
                print(f"\n   💡 Notes for this assignment type:")
                for note in notes[:3]:
                    print(f"      • {note}")
            
            print()
            confirm = input("Use this profile? (y/n, default=y): ").strip().lower()
            if confirm != 'n':
                # Ask about citation verification for paper types that might include sources
                global VERIFY_CITATIONS
                citation_profiles = ["research_paper", "analytical_essay", "personal_reflection", "discussion_post", "standard"]
                
                if profile_key in citation_profiles:
                    print()
                    print("   📚 Citation Verification Option")
                    print("   This feature checks if in-text citations exist in CrossRef/OpenLibrary.")
                    print("   Useful for detecting AI-fabricated sources.")
                    print("   ⚠️  Note: Requires internet access and adds ~0.5 seconds per citation.")
                    print()
                    
                    if profile_key == "research_paper":
                        default_verify = "y"
                        prompt = "   Verify citations against academic databases? (y/n, default=y): "
                    else:
                        default_verify = "n"
                        prompt = "   Verify citations against academic databases? (y/n, default=n): "
                    
                    verify = input(prompt).strip().lower()
                    if verify == "":
                        verify = default_verify
                    
                    VERIFY_CITATIONS = (verify == 'y')
                    if VERIFY_CITATIONS:
                        print("   ✅ Citation verification ENABLED")
                    else:
                        print("   ℹ️  Citation verification skipped (pattern checks still active)")
                else:
                    VERIFY_CITATIONS = False
                
                return profile_key
            else:
                print_profile_selection_menu()
        else:
            print("❌ Invalid choice. Please enter 1-8.")

def print_profile_summary(profile_key: str):
    """Print a summary of the selected profile for the report."""
    profile = get_profile(profile_key)
    
    print(f"\n📋 Analysis Profile: {profile['name']}")
    print(f"   {profile['description']}")
    
    expectations = profile.get("expectations", {})
    print(f"   Expected polish level: {expectations.get('polish_level', 'medium')}")
    print(f"   Personal voice: {expectations.get('personal_voice', 'varies')}")

# AI detection patterns - clichéd transitions and meta-commentary
AI_TRANSITIONS = [
    "it is important to note", "it should be noted", "it is worth noting",
    "in conclusion", "to sum up", "in summary", "to summarize",
    "furthermore", "moreover", "additionally", "in addition",
    "this essay will explore", "this paper will examine", "this section will discuss",
    "as previously mentioned", "as stated above", "as discussed earlier",
    "on the one hand", "on the other hand", "conversely", "however",
    "firstly", "secondly", "thirdly", "lastly", "finally"
]

# Hedge phrases and formal markers
HEDGE_PHRASES = [
    "it can be argued that", "it could be said that", "one might argue",
    "it is possible that", "arguably", "presumably", "ostensibly",
    "it appears that", "it seems that", "tends to suggest"
]

# Vocabulary inflation - unnecessarily complex words
INFLATED_VOCAB = [
    ("utilize", "use"), ("demonstrate", "show"), ("individuals", "people"),
    ("commence", "start"), ("terminate", "end"), ("endeavor", "try"),
    ("facilitate", "help"), ("implement", "do"), ("ascertain", "find out"),
    ("optimal", "best"), ("subsequent", "next"), ("prior to", "before"),
    ("in order to", "to"), ("due to the fact that", "because")
]

# Generic/vague phrases lacking specificity
GENERIC_PHRASES = [
    "many things", "various aspects", "in today's society", "in today's world",
    "throughout history", "since the beginning of time", "it can be said",
    "some people believe", "it is believed that", "one could argue",
    "studies show", "research indicates", "experts say", "it has been shown",
    "this shows that", "this proves that", "overall", "basically",
    "a variety of", "a number of", "plays a crucial role", "of paramount importance"
]

# Balanced/false equivalence markers
BALANCE_MARKERS = [
    "both sides", "both perspectives", "different viewpoints", "various opinions",
    "while some argue", "others contend", "there are arguments for and against"
]

# Passive voice patterns
PASSIVE_PATTERNS = [
    r"\b(is|are|was|were|be|been|being)\s+\w+ed\b",
    r"\b(can|could|may|might|should|would)\s+be\s+\w+ed\b"
]

# Personal/embodied markers (lack of these is a flag)
PERSONAL_MARKERS = [
    r"\bi\s", r"\bmy\b", r"\bme\b", r"\bmine\b",
    r"\bwe\b", r"\bour\b", r"\bus\b",
    r"as a\b", r"being a\b", r"growing up"
]

# Emotional/vulnerable language
EMOTIONAL_MARKERS = [
    "i feel", "i felt", "i was", "i struggled", "i realized", "i noticed",
    "confused", "frustrated", "excited", "surprised", "shocked", "angry",
    "it hurt", "i cried", "i laughed", "nervous", "scared", "proud"
]

# Generic filename patterns
GENERIC_FILENAMES = [
    r"^untitled\d*\..*$", r"^document\d*\..*$", r"^assignment\d*\..*$",
    r"^file\d*\..*$", r"^download\d*\..*$", r"^new file\d*\..*$",
    r"^copy of .*$"
]

def count_words(text: str) -> int:
    """Count words in text."""
    if not text:
        return 0
    return len(text.split())

def calculate_text_similarity(text1: str, text2: str) -> float:
    """Calculate simple similarity ratio between two texts."""
    if not text1 or not text2:
        return 0.0
    
    t1 = text1.lower().strip()
    t2 = text2.lower().strip()
    
    if t1 == t2:
        return 1.0
    
    words1 = set(t1.split())
    words2 = set(t2.split())
    
    if not words1 or not words2:
        return 0.0
    
    intersection = words1.intersection(words2)
    union = words1.union(words2)
    
    return len(intersection) / len(union) if union else 0.0

def get_all_discussion_topics(course_id: int) -> List[Dict]:
    """Fetch all discussion topics in the course."""
    print("💬 Fetching all discussion topics...")
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/discussion_topics"
    params = {"per_page": 100}
    response = requests.get(url, headers=HEADERS, params=params)

    if response.status_code != 200:
        print(f"❌ Failed to fetch discussion topics: {response.text}")
        return []

    data = response.json()
    return data if isinstance(data, list) else []


def fetch_discussion_entries(course_id: int, topic_id: int) -> List[Dict]:
    """Fetch all entries (including replies) for a discussion topic."""
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/discussion_topics/{topic_id}/view"
    resp = requests.get(url, headers=HEADERS)
    if resp.status_code != 200:
        print(f"⚠️  Failed to fetch discussion {topic_id} (HTTP {resp.status_code})")
        if "text/html" in resp.headers.get("content-type", ""):
            print("   💡 Likely cause: Canvas returned login page (check token/URL)")
        return []
    try:
        data = resp.json()
        entries = data.get("view", [])
    except Exception as e:
        print(f"   ❌ JSON decode error: {e}")
        return []
    return entries


def analyze_discussion_posts(course_id: int, topic: Dict, students: List[Dict]) -> Tuple[Dict[int, List[str]], Dict[int, str]]:
    """
    Analyze all posts in a discussion topic for academic dishonesty flags.
    Returns: (Dict mapping user_id to list of flags, Dict mapping user_id to post text)
    """
    topic_id = topic.get("id")
    topic_name = topic.get("title", f"Discussion {topic_id}")
    
    print(f"   💬 Analyzing discussion: {topic_name}")
    
    # Get all entries for this discussion
    entries = fetch_discussion_entries(course_id, topic_id)
    if not entries:
        return {}, {}
    
    # Flatten replies to get all posts
    all_posts = []
    def flatten_posts(posts):
        for post in posts:
            all_posts.append(post)
            if "replies" in post:
                flatten_posts(post["replies"])
    
    flatten_posts(entries)
    
    student_flags = {}
    post_texts = {}
    
    # Create a mapping of student user IDs for quick lookup
    student_user_ids = {student.get("user_id") for student in students if student.get("user_id")}
    
    for post in all_posts:
        user_id = post.get("user_id")
        
        # Only analyze posts from enrolled students
        if user_id not in student_user_ids:
            continue
        
        message = post.get("message", "")
        if message:
            # Create a submission-like object for compatibility with analyze_submission
            submission = {
                "body": message,
                "user_id": user_id,
                "workflow_state": "submitted"
            }
            
            flags = analyze_submission(submission, all_posts)
            if flags:
                if user_id not in student_flags:
                    student_flags[user_id] = []
                student_flags[user_id].extend(flags)
            
            # Store post text for cross-analysis
            if user_id not in post_texts:
                post_texts[user_id] = ""
            post_texts[user_id] += " " + message
    
    return student_flags, post_texts

def check_ai_transitions(text: str) -> int:
    """Count clichéd transitions and meta-commentary."""
    if not text:
        return 0
    
    text_lower = text.lower()
    count = 0
    
    for phrase in AI_TRANSITIONS:
        count += text_lower.count(phrase)
    
    return count

def check_hedge_phrases(text: str) -> int:
    """Count hedge phrases indicating avoidance of stance."""
    if not text:
        return 0
    
    text_lower = text.lower()
    count = 0
    
    for phrase in HEDGE_PHRASES:
        count += text_lower.count(phrase)
    
    return count

def check_inflated_vocabulary(text: str) -> List[str]:
    """Detect unnecessarily complex vocabulary."""
    if not text:
        return []
    
    text_lower = text.lower()
    found = []
    
    for complex_word, simple_word in INFLATED_VOCAB:
        if complex_word in text_lower:
            found.append(f"{complex_word} (vs {simple_word})")
    
    return found

def check_generic_content(text: str) -> int:
    """Count generic/vague phrases lacking specificity."""
    if not text:
        return 0
    
    text_lower = text.lower()
    count = 0
    
    for phrase in GENERIC_PHRASES:
        count += text_lower.count(phrase)
    
    return count

def check_balance_markers(text: str) -> int:
    """Check for false equivalence/over-balanced writing."""
    if not text:
        return 0
    
    text_lower = text.lower()
    count = 0
    
    for marker in BALANCE_MARKERS:
        count += text_lower.count(marker)
    
    return count

def check_passive_voice(text: str) -> int:
    """Count passive voice constructions."""
    if not text:
        return 0
    
    count = 0
    for pattern in PASSIVE_PATTERNS:
        count += len(re.findall(pattern, text, re.IGNORECASE))
    
    return count

def check_personal_markers(text: str) -> int:
    """Count first-person and embodied language."""
    if not text:
        return 0
    
    count = 0
    text_lower = text.lower()
    
    for pattern in PERSONAL_MARKERS:
        count += len(re.findall(pattern, text_lower))
    
    return count

def check_emotional_markers(text: str) -> int:
    """Count emotional/vulnerable language."""
    if not text:
        return 0
    
    text_lower = text.lower()
    count = 0
    
    for marker in EMOTIONAL_MARKERS:
        count += text_lower.count(marker)
    
    return count

def check_paragraph_uniformity(text: str) -> float:
    """Check if paragraphs are suspiciously uniform in length."""
    if not text:
        return 0.0
    
    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
    
    if len(paragraphs) < 3:
        return 0.0
    
    lengths = [len(p.split()) for p in paragraphs]
    avg_length = sum(lengths) / len(lengths)
    
    # Calculate coefficient of variation
    variance = sum((l - avg_length) ** 2 for l in lengths) / len(lengths)
    std_dev = variance ** 0.5
    
    if avg_length == 0:
        return 0.0
    
    coefficient_of_variation = std_dev / avg_length
    
    # Low variation = high uniformity (suspicious)
    return 1 - coefficient_of_variation if coefficient_of_variation < 1 else 0

def check_repetitive_reasoning(text: str) -> bool:
    """Detect repetitive or circular reasoning."""
    if not text:
        return False
    
    sentences = re.split(r'[.!?]+', text)
    sentences = [s.strip().lower() for s in sentences if len(s.strip()) > 20]
    
    if len(sentences) < 5:
        return False
    
    # Check for high similarity between sentences
    similar_pairs = 0
    for i in range(len(sentences) - 1):
        for j in range(i + 1, len(sentences)):
            similarity = calculate_text_similarity(sentences[i], sentences[j])
            if similarity > 0.7:
                similar_pairs += 1
    
    # If more than 20% of sentence pairs are highly similar
    total_pairs = (len(sentences) * (len(sentences) - 1)) / 2
    return (similar_pairs / total_pairs) > 0.2 if total_pairs > 0 else False

def check_copy_paste_indicators(text: str) -> List[str]:
    """Check for copy-paste indicators."""
    if not text:
        return []
    
    indicators = []
    
    if re.search(r'  +', text):
        indicators.append("Multiple consecutive spaces")
    
    if re.search(r'\n{3,}', text):
        indicators.append("Excessive line breaks")
    
    if '\t' in text and '    ' in text:
        indicators.append("Mixed tab/space formatting")
    
    if re.search(r'\x00|\x0c|\ufeff', text):
        indicators.append("Special formatting characters")
    
    return indicators

def check_sentence_completeness(text: str) -> float:
    """Calculate ratio of complete sentences."""
    if not text:
        return 0.0
    
    sentences = re.split(r'[.!?]+', text)
    sentences = [s.strip() for s in sentences if s.strip()]
    
    if not sentences:
        return 0.0
    
    complete_count = 0
    for sentence in sentences:
        words = sentence.split()
        if len(words) >= 8 and sentence[0].isupper():
            complete_count += 1
    
    return complete_count / len(sentences)

def check_generic_filename(filename: str) -> bool:
    """Check if filename matches generic patterns."""
    if not filename:
        return False
    
    filename_lower = filename.lower()
    
    for pattern in GENERIC_FILENAMES:
        if re.match(pattern, filename_lower):
            return True
    
    return False

# ======================
# ESSAY-LEVEL ORGANIZATION CHECKS
# ======================

def check_essay_organization(text: str) -> Dict[str, Any]:
    """
    Analyze essay-level organizational patterns that may indicate AI generation.
    Returns dict with various organizational flags.
    """
    if not text or len(text) < 200:
        return {"flags": [], "details": {}}
    
    flags = []
    details = {}
    
    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
    if len(paragraphs) < 2:
        # Try single newline split
        paragraphs = [p.strip() for p in text.split('\n') if p.strip() and len(p.strip()) > 50]
    
    details["paragraph_count"] = len(paragraphs)
    
    # Check 1: Formulaic introduction patterns
    if paragraphs:
        first_para = paragraphs[0].lower()
        intro_patterns = [
            r"this essay will (explore|examine|discuss|analyze|argue)",
            r"in this (essay|paper|analysis|discussion)",
            r"the purpose of this (essay|paper)",
            r"this (paper|essay) aims to",
            r"throughout this (essay|paper)",
            r"the following (essay|paper|analysis) will"
        ]
        for pattern in intro_patterns:
            if re.search(pattern, first_para):
                flags.append("Formulaic AI-style introduction")
                details["formulaic_intro"] = True
                break
    
    # Check 2: Formulaic conclusion patterns
    if len(paragraphs) > 1:
        last_para = paragraphs[-1].lower()
        conclusion_patterns = [
            r"in conclusion,?\s*(it is clear|we can see|this essay has)",
            r"to (sum up|summarize|conclude),?\s*(the|this|it)",
            r"all (things considered|in all)",
            r"in (summary|closing),?\s*(this|the|it)",
            r"as (this essay|we) have (shown|seen|demonstrated|discussed)",
            r"the (foregoing|preceding|above) (discussion|analysis|examination)"
        ]
        for pattern in conclusion_patterns:
            if re.search(pattern, last_para):
                flags.append("Formulaic AI-style conclusion")
                details["formulaic_conclusion"] = True
                break
    
    # Check 3: Perfect 5-paragraph essay structure
    if len(paragraphs) == 5:
        # Check if it follows classic structure
        word_counts = [len(p.split()) for p in paragraphs]
        # Intro and conclusion typically shorter than body paragraphs
        if (word_counts[0] < word_counts[2] and 
            word_counts[4] < word_counts[2] and
            abs(word_counts[1] - word_counts[2]) < 30 and
            abs(word_counts[2] - word_counts[3]) < 30):
            flags.append("Classic 5-paragraph essay structure (common in AI)")
            details["five_paragraph_structure"] = True
    
    # Check 4: Suspiciously uniform paragraph lengths
    if len(paragraphs) >= 3:
        word_counts = [len(p.split()) for p in paragraphs]
        avg_words = sum(word_counts) / len(word_counts)
        if avg_words > 30:  # Only check substantial paragraphs
            variance = sum((w - avg_words) ** 2 for w in word_counts) / len(word_counts)
            std_dev = variance ** 0.5
            cv = std_dev / avg_words if avg_words > 0 else 0
            
            if cv < 0.15:  # Very uniform
                flags.append(f"Suspiciously uniform paragraph lengths (CV: {cv:.2f})")
                details["uniform_paragraphs"] = True
    
    # Check 5: Each paragraph starts with topic sentence + transition
    if len(paragraphs) >= 3:
        transition_starts = 0
        topic_sentence_pattern = 0
        
        transition_starters = [
            "first", "second", "third", "additionally", "furthermore", 
            "moreover", "however", "in addition", "another", "finally",
            "next", "then", "also", "similarly", "consequently"
        ]
        
        for i, para in enumerate(paragraphs[1:-1], 1):  # Skip intro/conclusion
            first_sentence = para.split('.')[0].lower() if para else ""
            
            # Check for transition word start
            for starter in transition_starters:
                if first_sentence.startswith(starter):
                    transition_starts += 1
                    break
            
            # Check for topic sentence pattern (makes a claim)
            if re.match(r'^[a-z]+ (is|are|was|were|has|have|can|should|must|will)', first_sentence):
                topic_sentence_pattern += 1
        
        body_para_count = len(paragraphs) - 2
        if body_para_count > 0:
            if transition_starts / body_para_count > 0.6:
                flags.append("Most body paragraphs start with transitions (mechanical)")
                details["transition_heavy"] = True
            
            if topic_sentence_pattern / body_para_count > 0.8:
                flags.append("Every paragraph follows topic-sentence formula")
                details["formulaic_topic_sentences"] = True
    
    return {"flags": flags, "details": details}

def check_headings_structure(text: str) -> Dict[str, Any]:
    """
    Check for AI-typical heading patterns in the text.
    """
    if not text:
        return {"flags": [], "details": {}}
    
    flags = []
    details = {}
    
    # Look for markdown-style or common heading patterns
    heading_patterns = [
        r'^#{1,3}\s+.+$',                    # Markdown headings
        r'^[A-Z][^.!?]*:$',                  # Title case with colon
        r'^\*\*[^*]+\*\*$',                  # Bold as heading
        r'^[IVX]+\.\s+[A-Z]',                # Roman numerals
        r'^\d+\.\s+[A-Z][^.!?]{5,}$',        # Numbered sections
        r'^[A-Z][a-z]+(\s+[A-Z][a-z]+){1,5}$'  # Title Case Lines
    ]
    
    lines = text.split('\n')
    headings_found = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        for pattern in heading_patterns:
            if re.match(pattern, line, re.MULTILINE):
                headings_found.append(line)
                break
    
    details["headings_count"] = len(headings_found)
    details["headings"] = headings_found[:5]  # Store first 5
    
    if len(headings_found) >= 3:
        # Check for AI-typical heading styles
        ai_heading_patterns = [
            r'introduction',
            r'background',
            r'overview', 
            r'main\s*(body|points?|arguments?)',
            r'(key\s*)?(points?|findings?|takeaways?)',
            r'(analysis|discussion)',
            r'conclusion',
            r'(final\s*)?thoughts?',
            r'summary'
        ]
        
        ai_style_count = 0
        for heading in headings_found:
            heading_lower = heading.lower()
            for pattern in ai_heading_patterns:
                if re.search(pattern, heading_lower):
                    ai_style_count += 1
                    break
        
        if ai_style_count >= 2:
            flags.append(f"AI-typical heading structure ({ai_style_count} generic headings)")
            details["ai_style_headings"] = True
        
        # Check for numbered headings (common in AI)
        numbered = sum(1 for h in headings_found if re.match(r'^\d+\.?\s', h))
        if numbered >= 3:
            flags.append("Numbered section headings (common in AI output)")
            details["numbered_headings"] = True
    
    return {"flags": flags, "details": details}

def check_paragraph_as_prose_in_notes(text: str) -> Dict[str, Any]:
    """
    For notes/outlines: Check if submission is actually prose paragraphs
    when it should be informal notes with fragments.
    """
    if not text:
        return {"flags": [], "details": {}}
    
    flags = []
    details = {}
    
    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
    if not paragraphs:
        paragraphs = [p.strip() for p in text.split('\n') if p.strip()]
    
    # Check for prose indicators
    prose_indicators = 0
    
    # 1. Long continuous paragraphs (notes should be short/bulleted)
    long_paragraphs = sum(1 for p in paragraphs if len(p.split()) > 40)
    if long_paragraphs >= 2:
        prose_indicators += 1
        details["long_paragraphs"] = long_paragraphs
    
    # 2. Lack of bullet points or numbering
    bullet_patterns = r'^[\-\*\•\◦\▪]\s|^\d+[\.\)]\s|^[a-z][\.\)]\s'
    bulleted_lines = sum(1 for line in text.split('\n') if re.match(bullet_patterns, line.strip()))
    total_lines = len([l for l in text.split('\n') if l.strip()])
    
    if total_lines > 5 and bulleted_lines / total_lines < 0.2:
        prose_indicators += 1
        details["lacks_bullets"] = True
    
    # 3. Sentences connected with transitions (prose, not notes)
    transition_words = ["however", "therefore", "furthermore", "additionally", 
                       "moreover", "consequently", "thus", "hence"]
    transition_count = sum(text.lower().count(t) for t in transition_words)
    if transition_count >= 3:
        prose_indicators += 1
        details["has_transitions"] = transition_count
    
    # 4. High complete sentence ratio
    sentences = re.split(r'[.!?]+', text)
    sentences = [s.strip() for s in sentences if s.strip()]
    complete = sum(1 for s in sentences if len(s.split()) >= 8 and s[0].isupper())
    ratio = complete / len(sentences) if sentences else 0
    
    if ratio > 0.6:
        prose_indicators += 1
        details["complete_sentence_ratio"] = ratio
    
    if prose_indicators >= 2:
        flags.append(f"Submission appears to be prose, not notes/outline (indicators: {prose_indicators})")
    
    if prose_indicators >= 3:
        flags.append("Strong evidence: Polished prose submitted as 'notes'")
    
    return {"flags": flags, "details": details}

# ======================
# BIBLIOGRAPHIC CHECKS
# ======================

AI_CITATION_PATTERNS = {
    "fabricated_author_patterns": [
        r'\b(Smith|Johnson|Williams|Brown|Jones|Davis|Miller|Wilson)\s+et\s+al\.',  # Common surnames + et al
        r'\([A-Z][a-z]+,\s*\d{4}\)',  # Simple (Author, Year) that's very common
    ],
    "suspicious_journal_names": [
        r'journal of [a-z]+ studies',
        r'international journal of [a-z]+',
        r'american journal of [a-z]+',
        r'quarterly review of [a-z]+'
    ],
    "round_year_patterns": [
        r'\((?:19|20)(?:[0-9]0)\)',  # Years ending in 0
        r'\((?:2015|2018|2019|2020)\)'  # Commonly fabricated years
    ],
    "vague_citations": [
        r'according to (?:research|studies|experts|scientists)',
        r'studies (?:show|suggest|indicate|have shown)',
        r'research (?:shows|suggests|indicates|has shown)',
        r'experts (?:say|believe|argue|suggest)',
        r'it has been (?:shown|proven|demonstrated|found)'
    ]
}

def verify_citation_crossref(author: str, year: str, title: str = None) -> Dict[str, Any]:
    """
    Verify a citation exists using the CrossRef API.
    
    Args:
        author: Author last name
        year: Publication year
        title: Optional title for more precise matching
    
    Returns:
        Dict with verification results
    """
    import urllib.parse
    import time
    
    result = {
        "verified": False,
        "confidence": "unknown",
        "matches": [],
        "error": None
    }
    
    try:
        # Build query
        query_parts = [author]
        if title:
            query_parts.append(title)
        
        query = " ".join(query_parts)
        encoded_query = urllib.parse.quote(query)
        
        # CrossRef API (free, no key required for basic queries)
        url = f"https://api.crossref.org/works?query={encoded_query}&rows=5&filter=from-pub-date:{year},until-pub-date:{year}"
        
        # Add polite pool header (recommended by CrossRef)
        headers = {
            "User-Agent": "CanvasAutograder/1.0 (Academic integrity checking tool; mailto:instructor@example.edu)"
        }
        
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            items = data.get("message", {}).get("items", [])
            
            if items:
                # Check if any match our author/year
                for item in items:
                    item_authors = item.get("author", [])
                    author_names = [a.get("family", "").lower() for a in item_authors]
                    
                    if author.lower() in author_names:
                        result["verified"] = True
                        result["confidence"] = "high"
                        result["matches"].append({
                            "title": item.get("title", ["Unknown"])[0] if item.get("title") else "Unknown",
                            "authors": ", ".join([f"{a.get('given', '')} {a.get('family', '')}" for a in item_authors[:3]]),
                            "year": item.get("published-print", {}).get("date-parts", [[None]])[0][0] or 
                                   item.get("published-online", {}).get("date-parts", [[None]])[0][0],
                            "doi": item.get("DOI"),
                            "source": "CrossRef"
                        })
                
                if not result["verified"] and items:
                    # Found items in that year but author doesn't match
                    result["confidence"] = "low"
                    result["matches"] = [{"note": f"Found {len(items)} items from {year} but author '{author}' not found"}]
            else:
                result["confidence"] = "not_found"
        else:
            result["error"] = f"API returned status {response.status_code}"
            
    except requests.exceptions.Timeout:
        result["error"] = "Request timed out"
    except requests.exceptions.RequestException as e:
        result["error"] = str(e)
    except Exception as e:
        result["error"] = f"Unexpected error: {str(e)}"
    
    return result

def verify_citation_openlibrary(author: str, year: str, title: str = None) -> Dict[str, Any]:
    """
    Verify a citation exists using OpenLibrary API (good for books).
    
    Args:
        author: Author last name
        year: Publication year
        title: Optional title
    
    Returns:
        Dict with verification results
    """
    import urllib.parse
    
    result = {
        "verified": False,
        "confidence": "unknown",
        "matches": [],
        "error": None
    }
    
    try:
        # OpenLibrary search API
        query = f"author:{author}"
        if title:
            query += f" title:{title}"
        
        encoded_query = urllib.parse.quote(query)
        url = f"https://openlibrary.org/search.json?q={encoded_query}&limit=5"
        
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            docs = data.get("docs", [])
            
            for doc in docs:
                pub_year = doc.get("first_publish_year")
                if pub_year and abs(int(pub_year) - int(year)) <= 2:  # Allow 2-year variance
                    result["verified"] = True
                    result["confidence"] = "medium"
                    result["matches"].append({
                        "title": doc.get("title", "Unknown"),
                        "authors": ", ".join(doc.get("author_name", [])[:3]),
                        "year": pub_year,
                        "source": "OpenLibrary"
                    })
            
            if not result["verified"] and docs:
                result["confidence"] = "low"
                result["matches"] = [{"note": f"Found works by '{author}' but not from {year}"}]
        else:
            result["error"] = f"API returned status {response.status_code}"
            
    except Exception as e:
        result["error"] = str(e)
    
    return result

def verify_citations_batch(citations: List[Tuple[str, str]], progress_callback=None) -> Dict[str, Any]:
    """
    Verify multiple citations with rate limiting.
    
    Args:
        citations: List of (author, year) tuples
        progress_callback: Optional function to call with progress updates
    
    Returns:
        Dict with verification results for all citations
    """
    import time
    
    results = {
        "total": len(citations),
        "verified": 0,
        "not_found": 0,
        "errors": 0,
        "details": []
    }
    
    for i, (author, year) in enumerate(citations):
        if progress_callback:
            progress_callback(i + 1, len(citations), author, year)
        
        # Try CrossRef first (better for journal articles)
        crossref_result = verify_citation_crossref(author, year)
        
        if crossref_result["verified"]:
            results["verified"] += 1
            results["details"].append({
                "author": author,
                "year": year,
                "status": "verified",
                "source": "CrossRef",
                "matches": crossref_result["matches"]
            })
        else:
            # Try OpenLibrary (better for books)
            openlib_result = verify_citation_openlibrary(author, year)
            
            if openlib_result["verified"]:
                results["verified"] += 1
                results["details"].append({
                    "author": author,
                    "year": year,
                    "status": "verified",
                    "source": "OpenLibrary",
                    "matches": openlib_result["matches"]
                })
            elif crossref_result.get("error") or openlib_result.get("error"):
                results["errors"] += 1
                results["details"].append({
                    "author": author,
                    "year": year,
                    "status": "error",
                    "error": crossref_result.get("error") or openlib_result.get("error")
                })
            else:
                results["not_found"] += 1
                results["details"].append({
                    "author": author,
                    "year": year,
                    "status": "not_found",
                    "note": "Citation could not be verified in CrossRef or OpenLibrary"
                })
        
        # Rate limiting: wait between requests to be polite to APIs
        if i < len(citations) - 1:
            time.sleep(0.5)  # 500ms between requests
    
    return results

def check_bibliographic_markers(text: str, verify_citations: bool = False) -> Dict[str, Any]:
    """
    Check for common AI patterns in citations and references.
    
    Args:
        text: The text to analyze
        verify_citations: If True, attempt to verify citations exist (slower, requires network)
    
    Returns:
        Dict with flags and details
    """
    if not text:
        return {"flags": [], "details": {}, "citations_found": [], "verification": None}
    
    flags = []
    details = {}
    citations_found = []
    
    text_lower = text.lower()
    
    # 1. Check for vague citation language (no actual citation)
    vague_count = 0
    for pattern in AI_CITATION_PATTERNS["vague_citations"]:
        matches = re.findall(pattern, text_lower)
        vague_count += len(matches)
    
    if vague_count >= 3:
        flags.append(f"Vague citation language without specific sources ({vague_count} instances)")
        details["vague_citations"] = vague_count
    
    # 2. Extract actual citations
    # APA style: (Author, Year) or Author (Year)
    apa_citations = re.findall(r'\(([A-Z][a-z]+(?:\s+(?:&|and)\s+[A-Z][a-z]+)?(?:\s+et\s+al\.)?),?\s*(\d{4})\)', text)
    apa_citations += re.findall(r'([A-Z][a-z]+(?:\s+(?:&|and)\s+[A-Z][a-z]+)?(?:\s+et\s+al\.)?)\s*\((\d{4})\)', text)
    
    # MLA style: (Author page)
    mla_citations = re.findall(r'\(([A-Z][a-z]+)\s+\d+\)', text)
    
    all_citations = [(author.split()[0].replace(",", ""), year) for author, year in apa_citations]  # Get first author surname
    details["citation_count"] = len(all_citations)
    citations_found = all_citations[:10]  # Store first 10
    
    # 3. Check for suspicious patterns in citations
    if all_citations:
        # Check for round years (2010, 2020, etc.)
        round_years = sum(1 for _, year in all_citations if year.endswith('0'))
        if len(all_citations) >= 3 and round_years / len(all_citations) > 0.5:
            flags.append(f"Suspicious: Many citations from round years ({round_years}/{len(all_citations)})")
        
        # Check for very common/generic author names
        common_names = ["smith", "johnson", "williams", "brown", "jones", "davis", "miller"]
        generic_authors = sum(1 for author, _ in all_citations if author.lower() in common_names)
        if generic_authors >= 2:
            flags.append(f"Multiple citations with very common author surnames ({generic_authors})")
        
        # Check for citations clustered in same year
        years = [year for _, year in all_citations]
        from collections import Counter
        year_counts = Counter(years)
        most_common_year, count = year_counts.most_common(1)[0] if year_counts else (None, 0)
        if count >= 3 and count / len(all_citations) > 0.4:
            flags.append(f"Many citations clustered in {most_common_year} ({count} citations)")
    
    # 4. Check for missing reference list
    has_references_section = bool(re.search(
        r'(references|works?\s+cited|bibliography)\s*[\n:]', 
        text_lower
    ))
    
    if len(all_citations) >= 3 and not has_references_section:
        flags.append("In-text citations present but no References/Works Cited section")
        details["missing_references"] = True
    
    # 5. Verify citations if requested
    verification_results = None
    if verify_citations and all_citations:
        print("\n🔍 Verifying citations (this may take a moment)...")
        
        def progress(current, total, author, year):
            print(f"   Checking {current}/{total}: {author} ({year})...")
        
        # Deduplicate citations
        unique_citations = list(set(all_citations))
        
        verification_results = verify_citations_batch(unique_citations, progress_callback=progress)
        
        # Add flags based on verification
        if verification_results["not_found"] > 0:
            not_found_list = [d for d in verification_results["details"] if d["status"] == "not_found"]
            flags.append(f"⚠️ {verification_results['not_found']} citation(s) could not be verified:")
            for item in not_found_list[:3]:
                flags.append(f"   - {item['author']} ({item['year']})")
            if len(not_found_list) > 3:
                flags.append(f"   - ... and {len(not_found_list) - 3} more")
        
        if verification_results["verified"] > 0:
            details["verified_citations"] = verification_results["verified"]
        
        # Summary
        details["verification_summary"] = {
            "total_checked": verification_results["total"],
            "verified": verification_results["verified"],
            "not_found": verification_results["not_found"],
            "errors": verification_results["errors"]
        }
    
    return {
        "flags": flags, 
        "details": details, 
        "citations_found": citations_found,
        "verification": verification_results
    }

# ======================
# SEMANTIC SIMILARITY DETECTION
# ======================

def extract_key_phrases(text: str, min_words: int = 3, max_words: int = 8) -> List[str]:
    """
    Extract key phrases from text for similarity comparison.
    Focuses on noun phrases and meaningful chunks.
    """
    if not text:
        return []
    
    # Clean text
    text = re.sub(r'<[^>]+>', '', text)  # Remove HTML
    text = re.sub(r'\s+', ' ', text).strip()
    
    phrases = []
    sentences = re.split(r'[.!?]+', text)
    
    for sentence in sentences:
        words = sentence.strip().split()
        if len(words) < min_words:
            continue
        
        # Extract sliding windows of phrases
        for length in range(min_words, min(max_words + 1, len(words) + 1)):
            for i in range(len(words) - length + 1):
                phrase = ' '.join(words[i:i + length]).lower()
                # Filter out phrases that are mostly function words
                content_words = [w for w in phrase.split() if len(w) > 3]
                if len(content_words) >= min_words - 1:
                    phrases.append(phrase)
    
    return phrases

def normalize_phrase(phrase: str) -> str:
    """
    Normalize a phrase for comparison (handles word order variations).
    """
    # Remove punctuation
    phrase = re.sub(r'[^\w\s]', '', phrase.lower())
    # Sort words to handle reordering
    words = sorted(phrase.split())
    return ' '.join(words)

def check_semantic_similarity_across_submissions(
    submissions: Dict[int, str],
    student_names: Dict[int, str],
    similarity_threshold: float = 0.6
) -> List[Dict]:
    """
    Check for semantically similar phrases across different students' submissions.
    Detects similar ideas expressed with slight wording variations.
    
    Args:
        submissions: Dict mapping user_id to submission text
        student_names: Dict mapping user_id to student name
        similarity_threshold: Minimum similarity for flagging
    
    Returns:
        List of similarity matches
    """
    matches = []
    
    # Extract phrases from each submission
    student_phrases = {}
    for user_id, text in submissions.items():
        if text and len(text) > 100:
            phrases = extract_key_phrases(text)
            # Also create normalized versions
            normalized = {normalize_phrase(p): p for p in phrases}
            student_phrases[user_id] = {
                "phrases": set(phrases),
                "normalized": normalized
            }
    
    # Compare across students
    user_ids = list(student_phrases.keys())
    
    for i, user1 in enumerate(user_ids):
        for user2 in user_ids[i + 1:]:
            data1 = student_phrases[user1]
            data2 = student_phrases[user2]
            
            # Check for exact phrase matches
            exact_matches = data1["phrases"].intersection(data2["phrases"])
            
            # Check for normalized matches (same words, different order)
            norm_matches = set()
            for norm1, orig1 in data1["normalized"].items():
                for norm2, orig2 in data2["normalized"].items():
                    if norm1 == norm2 and orig1 != orig2:
                        norm_matches.add((orig1, orig2))
            
            # Check for high word overlap in phrases
            similar_phrases = []
            for p1 in list(data1["phrases"])[:100]:  # Limit for performance
                words1 = set(p1.split())
                for p2 in list(data2["phrases"])[:100]:
                    words2 = set(p2.split())
                    if len(words1) >= 4 and len(words2) >= 4:
                        overlap = len(words1.intersection(words2))
                        union = len(words1.union(words2))
                        similarity = overlap / union if union > 0 else 0
                        
                        if similarity >= similarity_threshold and p1 != p2:
                            similar_phrases.append({
                                "phrase1": p1,
                                "phrase2": p2,
                                "similarity": similarity
                            })
            
            if exact_matches or norm_matches or similar_phrases:
                matches.append({
                    "student1": student_names.get(user1, f"User {user1}"),
                    "student1_id": user1,
                    "student2": student_names.get(user2, f"User {user2}"),
                    "student2_id": user2,
                    "exact_matches": list(exact_matches)[:5],
                    "reordered_matches": list(norm_matches)[:5],
                    "similar_phrases": similar_phrases[:5],
                    "total_matches": len(exact_matches) + len(norm_matches) + len(similar_phrases)
                })
    
    return matches

# ======================
# WHITE TEXT DETECTION
# ======================

def check_white_text_keywords(text: str, keywords: List[str]) -> Dict[str, Any]:
    """
    Check if specified 'white text' keywords appear in the submission.
    Teachers can embed hidden instructions in white text that AI would follow
    but students reading the assignment wouldn't see.
    
    Args:
        text: The submission text
        keywords: List of keywords that should appear if student used AI on the prompt
    
    Returns:
        Dict with found/missing keywords and analysis
    """
    if not text or not keywords:
        return {"flags": [], "found": [], "missing": [], "score": 0}
    
    text_lower = text.lower()
    
    found = []
    missing = []
    
    for keyword in keywords:
        keyword_lower = keyword.lower().strip()
        if keyword_lower in text_lower:
            found.append(keyword)
        else:
            missing.append(keyword)
    
    flags = []
    score = len(found) / len(keywords) if keywords else 0
    
    if found:
        flags.append(f"White text keywords detected: {', '.join(found)}")
        
        if score >= 0.8:
            flags.append("HIGH CONFIDENCE: Most/all white text keywords present - likely AI-generated from prompt")
        elif score >= 0.5:
            flags.append("MEDIUM CONFIDENCE: Multiple white text keywords found")
    
    return {
        "flags": flags,
        "found": found,
        "missing": missing,
        "score": score,
        "detection_rate": f"{int(score * 100)}%"
    }

def run_white_text_analysis(course_id: int, assignment_id: int, keywords: List[str]) -> Dict[str, Any]:
    """
    Run white text analysis on all submissions for an assignment.
    
    Args:
        course_id: Canvas course ID
        assignment_id: Canvas assignment ID
        keywords: List of white text keywords to check
    
    Returns:
        Dict with analysis results for all students
    """
    results = {
        "assignment_id": assignment_id,
        "keywords_checked": keywords,
        "students_flagged": [],
        "summary": {}
    }
    
    # Fetch submissions
    submissions = get_submissions(course_id, assignment_id)
    
    flagged_count = 0
    checked_count = 0
    
    for user_id, submission in submissions.items():
        body = submission.get("body", "")
        if not body:
            continue
        
        checked_count += 1
        user_name = submission.get("user", {}).get("name", f"User {user_id}")
        
        analysis = check_white_text_keywords(body, keywords)
        
        if analysis["found"]:
            flagged_count += 1
            results["students_flagged"].append({
                "user_id": user_id,
                "name": user_name,
                "keywords_found": analysis["found"],
                "detection_rate": analysis["detection_rate"],
                "flags": analysis["flags"]
            })
    
    results["summary"] = {
        "students_checked": checked_count,
        "students_flagged": flagged_count,
        "flag_rate": f"{int(flagged_count / checked_count * 100)}%" if checked_count > 0 else "N/A"
    }
    
    return results

# ======================
# DRAFT COMPARISON
# ======================

def compare_drafts(draft1: str, draft2: str) -> Dict[str, Any]:
    """
    Compare two drafts (rough vs final) to analyze revision patterns.
    
    Args:
        draft1: Earlier draft text
        draft2: Later draft text
    
    Returns:
        Dict with comparison analysis
    """
    if not draft1 or not draft2:
        return {"flags": [], "details": {}, "revision_score": 0}
    
    flags = []
    details = {}
    
    # Basic metrics
    words1 = draft1.split()
    words2 = draft2.split()
    
    details["draft1_words"] = len(words1)
    details["draft2_words"] = len(words2)
    details["word_change"] = len(words2) - len(words1)
    details["word_change_percent"] = ((len(words2) - len(words1)) / len(words1) * 100) if words1 else 0
    
    # 1. Check for identical submission
    similarity = calculate_text_similarity(draft1, draft2)
    details["overall_similarity"] = similarity
    
    if similarity >= 0.98:
        flags.append("CRITICAL: Drafts are essentially identical - no revision occurred")
        return {"flags": flags, "details": details, "revision_score": 0}
    
    if similarity >= 0.95:
        flags.append("Drafts are nearly identical (>95% similar) - minimal revision")
    
    # 2. Check sentence-level changes
    sentences1 = set(s.strip().lower() for s in re.split(r'[.!?]+', draft1) if s.strip())
    sentences2 = set(s.strip().lower() for s in re.split(r'[.!?]+', draft2) if s.strip())
    
    unchanged = sentences1.intersection(sentences2)
    added = sentences2 - sentences1
    removed = sentences1 - sentences2
    
    details["sentences_unchanged"] = len(unchanged)
    details["sentences_added"] = len(added)
    details["sentences_removed"] = len(removed)
    
    total_sentences = len(sentences1.union(sentences2))
    change_ratio = (len(added) + len(removed)) / total_sentences if total_sentences > 0 else 0
    details["sentence_change_ratio"] = change_ratio
    
    # 3. Evaluate revision quality
    revision_score = 0
    
    # Score based on actual changes
    if change_ratio >= 0.3:
        revision_score += 30
        details["revision_level"] = "substantial"
    elif change_ratio >= 0.15:
        revision_score += 20
        details["revision_level"] = "moderate"
    elif change_ratio >= 0.05:
        revision_score += 10
        details["revision_level"] = "light"
    else:
        details["revision_level"] = "minimal"
        flags.append("Very few sentence-level changes between drafts")
    
    # Check for structural changes
    paras1 = [p.strip() for p in draft1.split('\n\n') if p.strip()]
    paras2 = [p.strip() for p in draft2.split('\n\n') if p.strip()]
    
    if len(paras1) != len(paras2):
        revision_score += 10
        details["paragraph_structure_changed"] = True
    
    # Check for word-level changes (indicates actual revision vs regeneration)
    words1_set = set(words1)
    words2_set = set(words2)
    
    new_words = words2_set - words1_set
    removed_words = words1_set - words2_set
    
    unique_new_words = len([w for w in new_words if len(w) > 4])
    details["new_vocabulary_words"] = unique_new_words
    
    if unique_new_words >= 20:
        revision_score += 20
    elif unique_new_words >= 10:
        revision_score += 10
    
    # 4. Check for suspicious patterns
    # Complete rewrite with similar structure (might indicate new AI generation)
    if similarity < 0.5 and len(paras1) == len(paras2):
        para_structure_sim = sum(
            1 for p1, p2 in zip(paras1, paras2) 
            if abs(len(p1.split()) - len(p2.split())) < 10
        ) / len(paras1)
        
        if para_structure_sim > 0.7:
            flags.append("Low text similarity but similar paragraph structure - possible AI regeneration")
            details["possible_regeneration"] = True
    
    # Draft 2 significantly shorter (might indicate different source)
    if len(words2) < len(words1) * 0.7:
        flags.append("Final draft significantly shorter than rough draft - unusual")
    
    # Calculate final revision score (0-100)
    revision_score = min(revision_score, 100)
    
    if revision_score < 20:
        flags.append("Insufficient revision between drafts")
    
    return {
        "flags": flags,
        "details": details,
        "revision_score": revision_score
    }

def analyze_submission(submission: Dict[str, Any], all_submissions: List[Dict[str, Any]]) -> List[str]:
    """
    Analyze submission and return list of flags.
    Uses the current assignment profile to determine which checks to run.
    Returns: list_of_flags
    """
    global CURRENT_PROFILE
    flags = []
    
    # Check text body
    body = submission.get("body", "")
    if body:
        word_count = count_words(body)
        min_words = get_threshold("min_word_count")
        
        # 1. Length check (always enabled)
        if word_count > 0 and word_count < min_words:
            flags.append(f"Very short text ({word_count} words)")
        
        # 2. AI Transitions check
        if is_check_enabled("ai_transitions"):
            transition_count = check_ai_transitions(body)
            threshold = get_threshold("ai_transition_count")
            if transition_count >= threshold:
                flags.append(f"Clichéd transitions ({transition_count} instances)")
        
        # 3. Hedge phrases (avoiding stance)
        if is_check_enabled("hedge_phrases"):
            hedge_count = check_hedge_phrases(body)
            if hedge_count >= 2:
                flags.append(f"Excessive hedging ({hedge_count} hedge phrases)")
        
        # 4. Vocabulary inflation
        if is_check_enabled("inflated_vocabulary"):
            inflated = check_inflated_vocabulary(body)
            if len(inflated) >= 3:
                flags.append(f"Inflated vocabulary: {', '.join(inflated[:3])}")
        
        # 5. Generic/vague content
        if is_check_enabled("generic_phrases"):
            generic_count = check_generic_content(body)
            if generic_count >= 3:
                flags.append(f"Generic/vague content ({generic_count} vague phrases)")
        
        # 6. Over-balanced writing
        if is_check_enabled("balance_markers"):
            balance_count = check_balance_markers(body)
            if balance_count >= 2:
                flags.append(f"Over-balanced/false equivalence ({balance_count} markers)")
        
        # 7. Excessive passive voice
        if is_check_enabled("passive_voice"):
            passive_count = check_passive_voice(body)
            total_sentences = len(re.split(r'[.!?]+', body))
            passive_threshold = get_threshold("passive_voice_percent") / 100
            if total_sentences > 5 and (passive_count / total_sentences) > passive_threshold:
                flags.append(f"Excessive passive voice ({int((passive_count/total_sentences)*100)}%)")
        
        # 8. Lack of personal markers
        if is_check_enabled("personal_markers"):
            personal_count = check_personal_markers(body)
            if word_count > 200 and personal_count < 3:
                flags.append("Lacks personal/embodied language")
        
        # 9. Lack of emotional markers
        if is_check_enabled("emotional_markers"):
            emotional_count = check_emotional_markers(body)
            if word_count > 200 and emotional_count == 0:
                flags.append("No emotional/vulnerable language")
        
        # 10. Uniform paragraph structure (only for more formal work)
        if is_check_enabled("complete_sentences"):
            uniformity = check_paragraph_uniformity(body)
            if uniformity > 0.8:
                flags.append("Suspiciously uniform paragraphs (mechanical structure)")
        
        # 11. Repetitive reasoning (always enabled - indicates AI regardless of type)
        if check_repetitive_reasoning(body):
            flags.append("Repetitive/circular reasoning detected")
        
        # 12. Copy-paste indicators (always important)
        if is_check_enabled("copy_paste"):
            copy_indicators = check_copy_paste_indicators(body)
            if copy_indicators:
                flags.append(f"Copy-paste indicators: {', '.join(copy_indicators)}")
        
        # 13. Suspiciously polished (complete sentences)
        if is_check_enabled("complete_sentences"):
            completeness = check_sentence_completeness(body)
            completeness_threshold = get_threshold("complete_sentence_percent") / 100
            
            # Check if we should INVERT this check (for notes/outlines)
            profile = get_profile(CURRENT_PROFILE) if CURRENT_PROFILE else {}
            invert_checks = profile.get("invert_checks", {})
            
            if invert_checks.get("complete_sentences"):
                # For notes: flag HIGH completeness as suspicious
                if completeness > completeness_threshold and word_count > 50:
                    flags.append(f"Too polished for notes ({int(completeness*100)}% complete sentences - expected fragments)")
            else:
                # Normal: flag high completeness as AI indicator
                if completeness > completeness_threshold and word_count > 100:
                    flags.append(f"Suspiciously polished ({int(completeness*100)}% complete sentences)")
        
        # 14. Essay-level organization checks
        if is_check_enabled("essay_organization"):
            org_results = check_essay_organization(body)
            flags.extend(org_results.get("flags", []))
        
        # 15. Heading structure checks
        if is_check_enabled("headings_structure"):
            heading_results = check_headings_structure(body)
            flags.extend(heading_results.get("flags", []))
        
        # 16. For notes/outlines: check if submission is actually prose
        if is_check_enabled("paragraph_structure"):
            profile = get_profile(CURRENT_PROFILE) if CURRENT_PROFILE else {}
            invert_checks = profile.get("invert_checks", {})
            
            if invert_checks.get("paragraph_structure"):
                # Check if what should be notes is actually prose
                prose_results = check_paragraph_as_prose_in_notes(body)
                flags.extend(prose_results.get("flags", []))
        
        # 17. Bibliographic checks (if text is long enough to likely have citations)
        if word_count > 300:
            bib_results = check_bibliographic_markers(body, verify_citations=VERIFY_CITATIONS)
            flags.extend(bib_results.get("flags", []))
        
        # 18. Check for cross-submission similarity
        if is_check_enabled("cross_submission"):
            current_user_id = submission.get("user_id")
            for other_sub in all_submissions:
                other_user_id = other_sub.get("user_id")
                if other_user_id == current_user_id:
                    continue
                
                other_body = other_sub.get("body", "")
                if other_body:
                    similarity = calculate_text_similarity(body, other_body)
                    if similarity >= DUPLICATE_SIMILARITY_THRESHOLD:
                        other_user_name = other_sub.get("user", {}).get("name", f"User {other_user_id}")
                        flags.append(f"Highly similar to {other_user_name}'s submission ({int(similarity*100)}% match)")
                        break
    
    # Check attachments (always enabled - applies to all types)
    attachments = submission.get("attachments", [])
    if attachments:
        for file in attachments:
            file_size = file.get("size", 0)
            file_name = file.get("filename", "unknown")
            
            if file_size > 0 and file_size < MIN_FILE_SIZE:
                flags.append(f"Small file '{file_name}' ({file_size} bytes)")
            
            if check_generic_filename(file_name):
                flags.append(f"Generic filename: '{file_name}'")
    
    url = submission.get("url")
    if url and not body and not attachments:
        flags.append("URL with no description")
    
    return flags

def get_active_students(course_id: int) -> List[Dict]:
    """Fetch active student enrollments."""
    print("📥 Fetching active student enrollments...")
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/enrollments"
    params = {"type": ["StudentEnrollment"], "state": ["active"], "per_page": 100}
    response = requests.get(url, headers=HEADERS, params=params)

    if not response.headers.get('Content-Type', '').startswith('application/json'):
        print("⚠️ Received non-JSON response")
        return []

    if response.status_code != 200:
        print(f"❌ Failed to fetch enrollments: {response.text}")
        return []

    data = response.json()
    return data if isinstance(data, list) else []

def get_all_assignments(course_id: int) -> List[Dict]:
    """Fetch all assignments in the course."""
    print("📚 Fetching all assignments...")
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/assignments"
    params = {"per_page": 100}
    response = requests.get(url, headers=HEADERS, params=params)

    if response.status_code != 200:
        print(f"❌ Failed to fetch assignments: {response.text}")
        return []

    data = response.json()
    return data if isinstance(data, list) else []

def get_submissions(course_id: int, assignment_id: int) -> Dict[int, Dict]:
    """Fetch submissions for an assignment."""
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions"
    params = {"include": ["attachments", "user"], "per_page": 100}
    response = requests.get(url, headers=HEADERS, params=params)

    if response.status_code != 200:
        return {}

    data = response.json()
    if not isinstance(data, list):
        return {}

    return {sub.get("user_id"): sub for sub in data if sub.get("user_id")}

def analyze_assignment(course_id: int, assignment: Dict, students: List[Dict]) -> Tuple[Dict[int, List[str]], Dict[int, str]]:
    """
    Analyze all submissions for an assignment.
    Returns: (Dict mapping user_id to list of flags, Dict mapping user_id to submission body)
    """
    assignment_id = assignment.get("id")
    assignment_name = assignment.get("name", f"Assignment {assignment_id}")
    
    print(f"   📝 Analyzing: {assignment_name}")
    
    submissions = get_submissions(course_id, assignment_id)
    if not submissions:
        return {}, {}
    
    all_submissions_list = list(submissions.values())
    student_flags = {}
    submission_texts = {}
    
    for student in students:
        user_id = student.get("user_id")
        if not user_id:
            continue
        
        submission = submissions.get(user_id)
        if submission and submission.get("workflow_state") != "unsubmitted":
            flags = analyze_submission(submission, all_submissions_list)
            if flags:
                student_flags[user_id] = flags
            
            # Store submission text for cross-assignment analysis
            body = submission.get("body", "")
            if body:
                submission_texts[user_id] = body
    
    return student_flags, submission_texts

def check_cross_assignment_similarity(all_data: Dict, student_names: Dict[int, str]) -> Tuple[Dict, Dict]:
    """
    Check for suspicious similarities across assignments.
    Returns: (student_self_plagiarism_dict, cross_student_plagiarism_dict)
    """
    print("\n🔍 Checking for cross-assignment similarities...")
    
    # Build structure: {user_id: {assignment_id: text}}
    student_submissions = defaultdict(dict)
    
    for assignment_id, assignment_data in all_data.items():
        for user_id, text in assignment_data.get("submission_texts", {}).items():
            if text and len(text.split()) > 50:  # Only check substantial submissions
                student_submissions[user_id][assignment_id] = text
    
    # Check 1: Self-plagiarism (student reusing their own work)
    self_plagiarism = {}
    
    for user_id, assignments in student_submissions.items():
        assignment_ids = list(assignments.keys())
        similar_pairs = []
        
        for i in range(len(assignment_ids)):
            for j in range(i + 1, len(assignment_ids)):
                aid1, aid2 = assignment_ids[i], assignment_ids[j]
                text1, text2 = assignments[aid1], assignments[aid2]
                
                similarity = calculate_text_similarity(text1, text2)
                if similarity >= 0.7:  # 70% threshold for self-plagiarism
                    assignment_name1 = all_data[aid1]["name"]
                    assignment_name2 = all_data[aid2]["name"]
                    similar_pairs.append({
                        "assignment1": assignment_name1,
                        "assignment1_id": aid1,
                        "assignment2": assignment_name2,
                        "assignment2_id": aid2,
                        "similarity": similarity
                    })
        
        if similar_pairs:
            self_plagiarism[user_id] = similar_pairs
    
    # Check 2: Cross-student plagiarism (different students, different assignments)
    cross_student_plagiarism = []
    
    # Build list of all (user_id, assignment_id, text) tuples
    all_submissions = []
    for user_id, assignments in student_submissions.items():
        for assignment_id, text in assignments.items():
            all_submissions.append((user_id, assignment_id, text))
    
    # Compare across students and assignments
    checked_pairs = set()
    
    for i in range(len(all_submissions)):
        user1, aid1, text1 = all_submissions[i]
        
        for j in range(i + 1, len(all_submissions)):
            user2, aid2, text2 = all_submissions[j]
            
            # Skip if same student or same assignment
            if user1 == user2 or aid1 == aid2:
                continue
            
            # Skip if we've already checked this pair
            pair_key = tuple(sorted([(user1, aid1), (user2, aid2)]))
            if pair_key in checked_pairs:
                continue
            checked_pairs.add(pair_key)
            
            similarity = calculate_text_similarity(text1, text2)
            if similarity >= 0.75:  # 75% threshold for cross-student
                cross_student_plagiarism.append({
                    "student1": student_names.get(user1, f"User {user1}"),
                    "student1_id": user1,
                    "assignment1": all_data[aid1]["name"],
                    "assignment1_id": aid1,
                    "student2": student_names.get(user2, f"User {user2}"),
                    "student2_id": user2,
                    "assignment2": all_data[aid2]["name"],
                    "assignment2_id": aid2,
                    "similarity": similarity
                })
    
    print(f"   Found {len(self_plagiarism)} students with self-plagiarism")
    print(f"   Found {len(cross_student_plagiarism)} cross-student similarities")
    
    return self_plagiarism, cross_student_plagiarism

def generate_report(course_id: int, course_name: str, all_data: Dict, self_plagiarism: Dict, cross_student_plagiarism: List, discussion_data: Dict = None):
    """Generate comprehensive report including assignments and discussion forums."""
    global CURRENT_PROFILE
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Get profile information for report
    profile = get_profile(CURRENT_PROFILE) if CURRENT_PROFILE else get_profile("standard")
    profile_name = profile.get("name", "Standard Analysis")
    
    # Cross-platform output base
    BASE_DIR = get_output_base_dir()

    # Academic Dishonesty outputs with CSV/Excel subfolders
    output_dir = BASE_DIR / "Academic Dishonesty Reports"
    csv_dir = output_dir / "csv"
    excel_dir = output_dir / "excel"
    
    try:
        csv_dir.mkdir(parents=True, exist_ok=True)
        excel_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        print(f"❌ Failed to create output directories: {e}")
        return
    
    # Build student summary data first (used for both console and file output)
    student_summary = defaultdict(lambda: {"total_flags": 0, "assignments": {}, "name": ""})
    
    for assignment_id, assignment_data in all_data.items():
        assignment_name = assignment_data["name"]
        for user_id, flags in assignment_data["flags"].items():
            student_name = assignment_data["student_names"].get(user_id, f"User {user_id}")
            student_summary[user_id]["name"] = student_name
            student_summary[user_id]["total_flags"] += len(flags)
            student_summary[user_id]["assignments"][assignment_name] = flags
    
    # Build discussion summary if available
    discussion_summary = defaultdict(lambda: {"total_flags": 0, "discussions": {}, "name": ""})
    if discussion_data:
        for topic_id, topic_data in discussion_data.items():
            topic_name = topic_data["name"]
            for user_id, flags in topic_data["flags"].items():
                student_name = topic_data["student_names"].get(user_id, f"User {user_id}")
                discussion_summary[user_id]["name"] = student_name
                discussion_summary[user_id]["total_flags"] += len(flags)
                discussion_summary[user_id]["discussions"][topic_name] = flags
    
    # Sort students by total flags (highest first)
    sorted_students = sorted(student_summary.items(), key=lambda x: x[1]["total_flags"], reverse=True)
    sorted_discussion_students = sorted(discussion_summary.items(), key=lambda x: x[1]["total_flags"], reverse=True)
    
    # =========================================================================
    # CONSOLE REPORT - Start with Executive Summary
    # =========================================================================
    print("\n" + "="*80)
    print(f"📊 ACADEMIC INTEGRITY FLAG REPORT: {course_name}")
    print("="*80)
    
    # Show analysis profile used
    print(f"\n📋 Analysis Profile: {profile_name}")
    expectations = profile.get("expectations", {})
    print(f"   Expected polish level: {expectations.get('polish_level', 'medium')}")
    print(f"   Personal voice requirement: {expectations.get('personal_voice', 'varies')}")
    
    # Show any instructor notes
    notes = profile.get("instructor_notes", [])
    if notes:
        print(f"\n   💡 Keep in mind:")
        for note in notes[:2]:
            print(f"      • {note}")
    
    # Executive Summary Table
    print("\n" + "="*80)
    print("📋 EXECUTIVE SUMMARY - Students Ranked by Flag Count")
    print("="*80)
    print()
    
    if sorted_students:
        # Calculate column widths
        max_name_len = max(len(data["name"]) for _, data in sorted_students)
        max_name_len = max(max_name_len, 12)  # Minimum width for "Student Name"
        
        # Print header
        print(f"{'Rank':<6} {'Student Name':<{max_name_len}} {'Flags':<8} {'Assignments':<12} {'Avg Flags/Assign'}")
        print("-" * (6 + max_name_len + 8 + 12 + 20))
        
        # Print each student
        for rank, (user_id, data) in enumerate(sorted_students, 1):
            num_assignments = len(data["assignments"])
            avg_flags = data["total_flags"] / num_assignments if num_assignments > 0 else 0
            
            # Add visual indicator for high-flag students
            if data["total_flags"] >= 10:
                indicator = "🚨"
            elif data["total_flags"] >= 5:
                indicator = "⚠️ "
            else:
                indicator = "   "
            
            print(f"{indicator}{rank:<3} {data['name']:<{max_name_len}} {data['total_flags']:<8} {num_assignments:<12} {avg_flags:.1f}")
        
        print()
        print(f"Legend: 🚨 = 10+ flags (high priority)  ⚠️ = 5-9 flags (review recommended)")
    else:
        print("   No flags found in assignments.")
    
    # Discussion summary if available
    if sorted_discussion_students:
        print("\n" + "-"*80)
        print("💬 Discussion Forum Flags Summary")
        print("-"*80)
        
        max_name_len = max(len(data["name"]) for _, data in sorted_discussion_students)
        max_name_len = max(max_name_len, 12)
        
        print(f"{'Rank':<6} {'Student Name':<{max_name_len}} {'Flags':<8} {'Discussions'}")
        print("-" * (6 + max_name_len + 8 + 15))
        
        for rank, (user_id, data) in enumerate(sorted_discussion_students, 1):
            print(f"{rank:<6} {data['name']:<{max_name_len}} {data['total_flags']:<8} {len(data['discussions'])}")
    
    # Detailed breakdown by student
    print("\n" + "="*80)
    print("👥 DETAILED FLAGS BY STUDENT (ASSIGNMENTS)")
    print("="*80)
    
    for user_id, data in sorted_students:
        print(f"\n👤 {data['name']} (ID: {user_id})")
        print(f"   Total flags: {data['total_flags']} across {len(data['assignments'])} assignments")
        print("-" * 80)
        for assignment_name, flags in data["assignments"].items():
            print(f"\n   📋 {assignment_name}:")
            for flag in flags:
                print(f"      • {flag}")
    
    # Add section for discussion forums if available
    if discussion_data and sorted_discussion_students:
        print("\n" + "="*80)
        print("💬 DETAILED FLAGS IN DISCUSSION FORUMS")
        print("="*80)
        
        for user_id, data in sorted_discussion_students:
            print(f"\n👤 {data['name']} (ID: {user_id})")
            print(f"   Total discussion flags: {data['total_flags']} across {len(data['discussions'])} discussions")
            print("-" * 80)
            for topic_name, flags in data["discussions"].items():
                print(f"\n   💬 {topic_name}:")
                for flag in flags:
                    print(f"      • {flag}")

    # CSV Export (detailed list for assignments)
    csv_path = csv_dir / f"canvas_flag_report_{course_id}_{timestamp}.csv"
    
    try:
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Type", "Student Name", "User ID", "Item", "Item ID", "Flag"])
            
            # Write assignment flags
            for assignment_id, assignment_data in all_data.items():
                assignment_name = assignment_data["name"]
                for user_id, flags in assignment_data["flags"].items():
                    student_name = assignment_data["student_names"].get(user_id, f"User {user_id}")
                    for flag in flags:
                        writer.writerow(["Assignment", student_name, user_id, assignment_name, assignment_id, flag])
            
            # Write discussion flags
            if discussion_data:
                for topic_id, topic_data in discussion_data.items():
                    topic_name = topic_data["name"]
                    for user_id, flags in topic_data["flags"].items():
                        student_name = topic_data["student_names"].get(user_id, f"User {user_id}")
                        for flag in flags:
                            writer.writerow(["Discussion", student_name, user_id, topic_name, topic_id, flag])
        
        print(f"\n✅ CSV report saved: {csv_path.name}")
            
    except Exception as e:
        print(f"❌ Error creating CSV file: {e}")
    
    # Excel Export with multiple sheets - Student-Centric Organization
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
        
        excel_path = excel_dir / f"canvas_flag_report_{course_id}_{timestamp}.xlsx"
        wb = openpyxl.Workbook()
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alert_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        warning_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # =====================================================
        # SHEET 1: Executive Summary (Student Rankings)
        # =====================================================
        ws_exec = wb.active
        ws_exec.title = "Executive Summary"
        
        exec_headers = ["Rank", "Student Name", "User ID", "Total Flags", "Assignments Flagged", 
                       "Avg Flags/Assignment", "Priority Level", "Top Flag Types"]
        for col, header in enumerate(exec_headers, 1):
            cell = ws_exec.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        row = 2
        for rank, (user_id, data) in enumerate(sorted_students, 1):
            num_assignments = len(data["assignments"])
            avg_flags = data["total_flags"] / num_assignments if num_assignments > 0 else 0
            
            # Determine priority level
            if data["total_flags"] >= 10:
                priority = "🚨 HIGH"
                row_fill = alert_fill
            elif data["total_flags"] >= 5:
                priority = "⚠️ MEDIUM"
                row_fill = warning_fill
            else:
                priority = "LOW"
                row_fill = None
            
            # Get top flag types
            all_flags = []
            for flags in data["assignments"].values():
                all_flags.extend(flags)
            
            # Count flag types
            flag_types = {}
            for flag in all_flags:
                flag_type = flag.split('(')[0].split(':')[0].strip()
                flag_types[flag_type] = flag_types.get(flag_type, 0) + 1
            
            top_flags = sorted(flag_types.items(), key=lambda x: x[1], reverse=True)[:3]
            top_flags_str = ", ".join([f"{t}({c})" for t, c in top_flags])
            
            ws_exec.cell(row=row, column=1, value=rank)
            ws_exec.cell(row=row, column=2, value=data["name"])
            ws_exec.cell(row=row, column=3, value=user_id)
            ws_exec.cell(row=row, column=4, value=data["total_flags"])
            ws_exec.cell(row=row, column=5, value=num_assignments)
            ws_exec.cell(row=row, column=6, value=round(avg_flags, 1))
            ws_exec.cell(row=row, column=7, value=priority)
            ws_exec.cell(row=row, column=8, value=top_flags_str)
            
            # Apply styling
            for col in range(1, 9):
                cell = ws_exec.cell(row=row, column=col)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
            
            row += 1
        
        # Adjust column widths
        ws_exec.column_dimensions['A'].width = 8
        ws_exec.column_dimensions['B'].width = 25
        ws_exec.column_dimensions['C'].width = 12
        ws_exec.column_dimensions['D'].width = 12
        ws_exec.column_dimensions['E'].width = 18
        ws_exec.column_dimensions['F'].width = 18
        ws_exec.column_dimensions['G'].width = 15
        ws_exec.column_dimensions['H'].width = 50
        
        # =====================================================
        # SHEET 2: Detailed Student View
        # =====================================================
        ws_students = wb.create_sheet("Student Details")
        
        student_headers = ["Student Name", "User ID", "Assignment", "Assignment ID", 
                         "Flag Count", "Flags"]
        for col, header in enumerate(student_headers, 1):
            cell = ws_students.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        row = 2
        for user_id, data in sorted_students:
            for assignment_name, flags in data["assignments"].items():
                # Find assignment ID
                assignment_id = None
                for aid, adata in all_data.items():
                    if adata["name"] == assignment_name:
                        assignment_id = aid
                        break
                
                ws_students.cell(row=row, column=1, value=data["name"])
                ws_students.cell(row=row, column=2, value=user_id)
                ws_students.cell(row=row, column=3, value=assignment_name)
                ws_students.cell(row=row, column=4, value=assignment_id)
                ws_students.cell(row=row, column=5, value=len(flags))
                ws_students.cell(row=row, column=6, value="; ".join(flags))
                
                for col in range(1, 7):
                    ws_students.cell(row=row, column=col).border = thin_border
                
                row += 1
        
        # Adjust widths
        ws_students.column_dimensions['A'].width = 25
        ws_students.column_dimensions['B'].width = 12
        ws_students.column_dimensions['C'].width = 30
        ws_students.column_dimensions['D'].width = 15
        ws_students.column_dimensions['E'].width = 12
        ws_students.column_dimensions['F'].width = 80
        
        # =====================================================
        # SHEET 3: Flag Type Analysis
        # =====================================================
        ws_flags = wb.create_sheet("Flag Analysis")
        
        # Aggregate all flags by type
        all_flag_types = defaultdict(lambda: {"count": 0, "students": set(), "assignments": set()})
        
        for assignment_id, assignment_data in all_data.items():
            for user_id, flags in assignment_data["flags"].items():
                for flag in flags:
                    flag_type = flag.split('(')[0].split(':')[0].strip()
                    all_flag_types[flag_type]["count"] += 1
                    all_flag_types[flag_type]["students"].add(user_id)
                    all_flag_types[flag_type]["assignments"].add(assignment_data["name"])
        
        flag_headers = ["Flag Type", "Total Occurrences", "Unique Students", "Assignments Affected"]
        for col, header in enumerate(flag_headers, 1):
            cell = ws_flags.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        row = 2
        for flag_type, data in sorted(all_flag_types.items(), key=lambda x: x[1]["count"], reverse=True):
            ws_flags.cell(row=row, column=1, value=flag_type)
            ws_flags.cell(row=row, column=2, value=data["count"])
            ws_flags.cell(row=row, column=3, value=len(data["students"]))
            ws_flags.cell(row=row, column=4, value=len(data["assignments"]))
            
            for col in range(1, 5):
                ws_flags.cell(row=row, column=col).border = thin_border
            
            row += 1
        
        ws_flags.column_dimensions['A'].width = 40
        ws_flags.column_dimensions['B'].width = 20
        ws_flags.column_dimensions['C'].width = 18
        ws_flags.column_dimensions['D'].width = 22
        
        # =====================================================
        # SHEET 4: Assignment Overview
        # =====================================================
        ws_assign = wb.create_sheet("Assignment Overview")
        
        assign_headers = ["Assignment Name", "Assignment ID", "Students Flagged", 
                        "Total Flags", "Most Common Flag"]
        for col, header in enumerate(assign_headers, 1):
            cell = ws_assign.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        row = 2
        for assignment_id, assignment_data in all_data.items():
            if not assignment_data["flags"]:
                continue
            
            total_flags = sum(len(f) for f in assignment_data["flags"].values())
            
            # Find most common flag
            flag_counts = defaultdict(int)
            for flags in assignment_data["flags"].values():
                for flag in flags:
                    flag_type = flag.split('(')[0].split(':')[0].strip()
                    flag_counts[flag_type] += 1
            
            most_common = max(flag_counts.items(), key=lambda x: x[1])[0] if flag_counts else "N/A"
            
            ws_assign.cell(row=row, column=1, value=assignment_data["name"])
            ws_assign.cell(row=row, column=2, value=assignment_id)
            ws_assign.cell(row=row, column=3, value=len(assignment_data["flags"]))
            ws_assign.cell(row=row, column=4, value=total_flags)
            ws_assign.cell(row=row, column=5, value=most_common)
            
            for col in range(1, 6):
                ws_assign.cell(row=row, column=col).border = thin_border
            
            row += 1
        
        ws_assign.column_dimensions['A'].width = 35
        ws_assign.column_dimensions['B'].width = 15
        ws_assign.column_dimensions['C'].width = 18
        ws_assign.column_dimensions['D'].width = 12
        ws_assign.column_dimensions['E'].width = 30
        
        # =====================================================
        # SHEET 5: Analysis Settings
        # =====================================================
        ws_settings = wb.create_sheet("Analysis Settings")
        
        ws_settings.cell(row=1, column=1, value="Analysis Profile Used")
        ws_settings.cell(row=1, column=2, value=profile_name)
        ws_settings.cell(row=2, column=1, value="Course Name")
        ws_settings.cell(row=2, column=2, value=course_name)
        ws_settings.cell(row=3, column=1, value="Analysis Date")
        ws_settings.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws_settings.cell(row=4, column=1, value="Total Assignments Analyzed")
        ws_settings.cell(row=4, column=2, value=len(all_data))
        ws_settings.cell(row=5, column=1, value="Total Students with Flags")
        ws_settings.cell(row=5, column=2, value=len(student_summary))
        
        # Add profile notes
        ws_settings.cell(row=7, column=1, value="Profile Notes:")
        ws_settings.cell(row=7, column=1).font = Font(bold=True)
        
        notes = profile.get("instructor_notes", [])
        for i, note in enumerate(notes, 8):
            ws_settings.cell(row=i, column=1, value=f"• {note}")
        
        ws_settings.column_dimensions['A'].width = 30
        ws_settings.column_dimensions['B'].width = 50
        
        wb.save(excel_path)
        print(f"✅ Excel report saved: {excel_path.name}")
        print(f"   Sheets: Executive Summary, Student Details, Flag Analysis, Assignment Overview, Settings")
        
    except ImportError:
        print("⚠️ openpyxl not installed. Excel export skipped.")
        print("   Install with: pip install openpyxl")
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
    

    
    # Summary statistics
    print("\n" + "="*80)
    print("📈 SUMMARY STATISTICS")
    print("="*80)
    
    total_assignments = len(all_data)
    total_flagged_assignments = sum(1 for a in all_data.values() if a["flags"])
    total_flags = sum(len(a["flags"]) for a in all_data.values())
    total_students_flagged = len(student_summary)
    
    print(f"Total assignments analyzed: {total_assignments}")
    print(f"Assignments with flags: {total_flagged_assignments}")
    print(f"Total flags issued: {total_flags}")
    print(f"Students with flags: {total_students_flagged}")
    
    if total_students_flagged > 0:
        avg_flags_per_student = sum(s["total_flags"] for s in student_summary.values()) / total_students_flagged
        print(f"Average flags per flagged student: {avg_flags_per_student:.1f}")

    # Final output location reminder
    print(f"\n📁 Reports were exported to:")
    print(f"   CSV: {csv_path}")
    if 'excel_path' in locals():
        print(f"   Excel: {excel_path}")


# ======================
# SPECIAL ANALYSIS MODES
# ======================

def run_white_text_mode(course_id: int, course_name: str):
    """
    Run white text keyword detection as a standalone quick analysis.
    """
    print("\n" + "="*70)
    print("🔑 WHITE TEXT DETECTION MODE")
    print("="*70)
    print()
    print("This tool checks if students' submissions contain specific keywords")
    print("that you embedded as white/hidden text in your assignment prompt.")
    print()
    print("If a student copies the prompt into an AI, those hidden keywords")
    print("may appear in the AI-generated response.")
    print()
    
    # Get assignment
    assignments = get_all_assignments(course_id)
    if not assignments:
        print("❌ No assignments found.")
        return
    
    print("Available assignments:")
    for i, a in enumerate(assignments[:20], 1):
        print(f"   [{i}] {a.get('name', 'Unnamed')} (ID: {a.get('id')})")
    
    if len(assignments) > 20:
        print(f"   ... and {len(assignments) - 20} more")
    
    try:
        assign_choice = input("\nEnter assignment number or ID: ").strip()
        if assign_choice.isdigit() and int(assign_choice) <= len(assignments):
            assignment = assignments[int(assign_choice) - 1]
        else:
            assignment = next((a for a in assignments if str(a.get('id')) == assign_choice), None)
        
        if not assignment:
            print("❌ Assignment not found.")
            return
    except (ValueError, IndexError):
        print("❌ Invalid selection.")
        return
    
    assignment_id = assignment.get('id')
    assignment_name = assignment.get('name', f'Assignment {assignment_id}')
    
    print(f"\n✅ Selected: {assignment_name}")
    
    # Get keywords
    print("\n" + "-"*50)
    print("Enter the white text keywords you embedded in the prompt.")
    print("These are words/phrases that should ONLY appear if a student")
    print("fed your prompt directly to an AI.")
    print()
    print("Enter keywords one per line. Press Enter twice when done:")
    print("-"*50)
    
    keywords = []
    while True:
        keyword = input("Keyword: ").strip()
        if not keyword:
            if keywords:
                break
            print("   (Enter at least one keyword)")
            continue
        keywords.append(keyword)
    
    print(f"\n🔍 Checking for {len(keywords)} keywords in submissions...")
    
    # Fetch submissions
    students = get_active_students(course_id)
    student_names = {s.get("user_id"): s.get("user", {}).get("name", f"User {s.get('user_id')}") for s in students}
    
    submissions_url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions"
    params = {"per_page": 100, "include[]": ["user"]}
    resp = requests.get(submissions_url, headers=HEADERS, params=params)
    
    if resp.status_code != 200:
        print(f"❌ Failed to fetch submissions: {resp.status_code}")
        return
    
    submissions = resp.json()
    
    # Analyze each submission
    results = []
    for sub in submissions:
        body = sub.get("body", "")
        if not body:
            continue
        
        user_id = sub.get("user_id")
        user_name = sub.get("user", {}).get("name", student_names.get(user_id, f"User {user_id}"))
        
        analysis = check_white_text_keywords(body, keywords)
        
        if analysis["found"]:
            results.append({
                "name": user_name,
                "user_id": user_id,
                "found": analysis["found"],
                "missing": analysis["missing"],
                "score": analysis["score"]
            })
    
    # Report results
    print("\n" + "="*70)
    print("📊 WHITE TEXT DETECTION RESULTS")
    print("="*70)
    print(f"\nAssignment: {assignment_name}")
    print(f"Keywords checked: {', '.join(keywords)}")
    print(f"Submissions analyzed: {len([s for s in submissions if s.get('body')])}")
    print()
    
    if results:
        print(f"🚨 FLAGGED STUDENTS: {len(results)}")
        print("-"*70)
        
        # Sort by detection rate
        results.sort(key=lambda x: x["score"], reverse=True)
        
        for r in results:
            print(f"\n👤 {r['name']} (ID: {r['user_id']})")
            print(f"   Detection rate: {int(r['score']*100)}%")
            print(f"   Keywords found: {', '.join(r['found'])}")
            if r['missing']:
                print(f"   Keywords missing: {', '.join(r['missing'])}")
            
            if r['score'] >= 0.8:
                print("   ⚠️  HIGH CONFIDENCE - Most keywords present")
            elif r['score'] >= 0.5:
                print("   ⚠️  MEDIUM CONFIDENCE - Multiple keywords present")
    else:
        print("✅ No students flagged - no keywords detected in any submission.")
    
    # Export option
    print()
    export = input("Export results to CSV? (y/n, default=n): ").strip().lower()
    if export == 'y':
        BASE_DIR = get_output_base_dir()
        output_dir = BASE_DIR / "Academic Dishonesty Reports" / "csv"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = output_dir / f"white_text_report_{course_id}_{assignment_id}_{timestamp}.csv"
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Student Name", "User ID", "Detection Rate", "Keywords Found", "Keywords Missing"])
            for r in results:
                writer.writerow([
                    r["name"], r["user_id"], f"{int(r['score']*100)}%",
                    ", ".join(r["found"]), ", ".join(r["missing"])
                ])
        
        print(f"✅ Exported to: {csv_path}")


def run_draft_comparison_mode(course_id: int, course_name: str):
    """
    Compare rough draft and final draft submissions to analyze revision patterns.
    """
    print("\n" + "="*70)
    print("📝 DRAFT COMPARISON MODE")
    print("="*70)
    print()
    print("This tool compares two versions of an assignment (e.g., rough draft")
    print("and final draft) to identify:")
    print("  • Students who submitted identical or near-identical drafts")
    print("  • Insufficient revision between versions")
    print("  • Possible AI regeneration (low similarity but same structure)")
    print()
    
    # Get assignments
    assignments = get_all_assignments(course_id)
    if not assignments:
        print("❌ No assignments found.")
        return
    
    print("Available assignments:")
    for i, a in enumerate(assignments[:20], 1):
        print(f"   [{i}] {a.get('name', 'Unnamed')} (ID: {a.get('id')})")
    
    if len(assignments) > 20:
        print(f"   ... and {len(assignments) - 20} more")
    
    # Select draft 1
    print("\n📄 Select ROUGH DRAFT (first/earlier version):")
    try:
        draft1_choice = input("Enter assignment number or ID: ").strip()
        if draft1_choice.isdigit() and int(draft1_choice) <= len(assignments):
            draft1_assignment = assignments[int(draft1_choice) - 1]
        else:
            draft1_assignment = next((a for a in assignments if str(a.get('id')) == draft1_choice), None)
        
        if not draft1_assignment:
            print("❌ Assignment not found.")
            return
    except (ValueError, IndexError):
        print("❌ Invalid selection.")
        return
    
    print(f"   ✅ Draft 1: {draft1_assignment.get('name')}")
    
    # Select draft 2
    print("\n📄 Select FINAL DRAFT (second/later version):")
    try:
        draft2_choice = input("Enter assignment number or ID: ").strip()
        if draft2_choice.isdigit() and int(draft2_choice) <= len(assignments):
            draft2_assignment = assignments[int(draft2_choice) - 1]
        else:
            draft2_assignment = next((a for a in assignments if str(a.get('id')) == draft2_choice), None)
        
        if not draft2_assignment:
            print("❌ Assignment not found.")
            return
    except (ValueError, IndexError):
        print("❌ Invalid selection.")
        return
    
    print(f"   ✅ Draft 2: {draft2_assignment.get('name')}")
    
    # Fetch submissions for both
    print("\n🔍 Fetching submissions...")
    
    def fetch_submissions(assignment_id):
        url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions"
        params = {"per_page": 100, "include[]": ["user"]}
        resp = requests.get(url, headers=HEADERS, params=params)
        if resp.status_code == 200:
            return {s.get("user_id"): s for s in resp.json()}
        return {}
    
    draft1_subs = fetch_submissions(draft1_assignment.get('id'))
    draft2_subs = fetch_submissions(draft2_assignment.get('id'))
    
    print(f"   Draft 1 submissions: {len(draft1_subs)}")
    print(f"   Draft 2 submissions: {len(draft2_subs)}")
    
    # Find students with both submissions
    common_students = set(draft1_subs.keys()).intersection(set(draft2_subs.keys()))
    print(f"   Students with both: {len(common_students)}")
    
    if not common_students:
        print("\n❌ No students have submitted both drafts.")
        return
    
    # Compare drafts
    print("\n🔍 Comparing drafts...")
    
    results = []
    for user_id in common_students:
        sub1 = draft1_subs[user_id]
        sub2 = draft2_subs[user_id]
        
        body1 = sub1.get("body", "")
        body2 = sub2.get("body", "")
        
        if not body1 or not body2:
            continue
        
        user_name = sub2.get("user", {}).get("name", f"User {user_id}")
        
        comparison = compare_drafts(body1, body2)
        
        results.append({
            "name": user_name,
            "user_id": user_id,
            "similarity": comparison["details"].get("overall_similarity", 0),
            "revision_score": comparison["revision_score"],
            "revision_level": comparison["details"].get("revision_level", "unknown"),
            "flags": comparison["flags"],
            "details": comparison["details"]
        })
    
    # Sort by revision score (lowest = most problematic)
    results.sort(key=lambda x: x["revision_score"])
    
    # Report
    print("\n" + "="*70)
    print("📊 DRAFT COMPARISON RESULTS")
    print("="*70)
    print(f"\nDraft 1: {draft1_assignment.get('name')}")
    print(f"Draft 2: {draft2_assignment.get('name')}")
    print(f"Students compared: {len(results)}")
    print()
    
    # Categorize results
    identical = [r for r in results if r["similarity"] >= 0.98]
    minimal_revision = [r for r in results if 0.95 <= r["similarity"] < 0.98]
    low_revision = [r for r in results if r["revision_score"] < 30 and r["similarity"] < 0.95]
    
    if identical:
        print(f"🚨 IDENTICAL SUBMISSIONS (>98% similar): {len(identical)}")
        for r in identical:
            print(f"   • {r['name']}: {int(r['similarity']*100)}% similar")
    
    if minimal_revision:
        print(f"\n⚠️  MINIMAL REVISION (95-98% similar): {len(minimal_revision)}")
        for r in minimal_revision:
            print(f"   • {r['name']}: {int(r['similarity']*100)}% similar")
    
    if low_revision:
        print(f"\n⚠️  LOW REVISION SCORE (<30): {len(low_revision)}")
        for r in low_revision:
            print(f"   • {r['name']}: Score {r['revision_score']}, {r['revision_level']} revision")
    
    # Show all results
    print("\n" + "-"*70)
    print("ALL STUDENTS (sorted by revision score, lowest first):")
    print("-"*70)
    print(f"{'Student':<25} {'Similarity':<12} {'Rev. Score':<12} {'Level':<15} {'Flags'}")
    print("-"*70)
    
    for r in results:
        flag_summary = str(len(r['flags'])) + " flags" if r['flags'] else "OK"
        print(f"{r['name'][:24]:<25} {int(r['similarity']*100)}%{'':<9} {r['revision_score']:<12} {r['revision_level']:<15} {flag_summary}")
    
    # Export option
    print()
    export = input("Export detailed results to CSV? (y/n, default=n): ").strip().lower()
    if export == 'y':
        BASE_DIR = get_output_base_dir()
        output_dir = BASE_DIR / "Academic Dishonesty Reports" / "csv"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = output_dir / f"draft_comparison_{course_id}_{timestamp}.csv"
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Student Name", "User ID", "Similarity %", "Revision Score", 
                "Revision Level", "Flags", "Word Change", "Sentences Added", "Sentences Removed"
            ])
            for r in results:
                writer.writerow([
                    r["name"], r["user_id"], f"{int(r['similarity']*100)}%",
                    r["revision_score"], r["revision_level"],
                    "; ".join(r["flags"]) if r["flags"] else "None",
                    r["details"].get("word_change", "N/A"),
                    r["details"].get("sentences_added", "N/A"),
                    r["details"].get("sentences_removed", "N/A")
                ])
        
        print(f"✅ Exported to: {csv_path}")


def run_phrase_similarity_mode(course_id: int, course_name: str):
    """
    Analyze submissions for semantically similar phrases across different students.
    Detects when multiple students may have used the same AI prompt.
    """
    print("\n" + "="*70)
    print("🔗 CROSS-STUDENT PHRASE SIMILARITY ANALYSIS")
    print("="*70)
    print()
    print("This tool identifies similar phrases across different students'")
    print("submissions, which may indicate:")
    print("  • Students using the same AI prompt")
    print("  • Shared AI-generated content")
    print("  • Similar phrasing with slight word variations")
    print()
    print("⚠️  Note: This analysis may take a while for large classes.")
    print()
    
    # Get assignment
    assignments = get_all_assignments(course_id)
    if not assignments:
        print("❌ No assignments found.")
        return
    
    print("Available assignments:")
    for i, a in enumerate(assignments[:20], 1):
        print(f"   [{i}] {a.get('name', 'Unnamed')} (ID: {a.get('id')})")
    
    try:
        assign_choice = input("\nEnter assignment number or ID: ").strip()
        if assign_choice.isdigit() and int(assign_choice) <= len(assignments):
            assignment = assignments[int(assign_choice) - 1]
        else:
            assignment = next((a for a in assignments if str(a.get('id')) == assign_choice), None)
        
        if not assignment:
            print("❌ Assignment not found.")
            return
    except (ValueError, IndexError):
        print("❌ Invalid selection.")
        return
    
    assignment_id = assignment.get('id')
    assignment_name = assignment.get('name', f'Assignment {assignment_id}')
    
    print(f"\n✅ Selected: {assignment_name}")
    
    # Fetch submissions
    print("\n🔍 Fetching submissions...")
    
    students = get_active_students(course_id)
    student_names = {s.get("user_id"): s.get("user", {}).get("name", f"User {s.get('user_id')}") for s in students}
    
    url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions"
    params = {"per_page": 100, "include[]": ["user"]}
    resp = requests.get(url, headers=HEADERS, params=params)
    
    if resp.status_code != 200:
        print(f"❌ Failed to fetch submissions: {resp.status_code}")
        return
    
    submissions_raw = resp.json()
    
    # Build submissions dict
    submissions = {}
    for sub in submissions_raw:
        user_id = sub.get("user_id")
        body = sub.get("body", "")
        if body and len(body) > 100:
            # Update student names from submission data
            if sub.get("user", {}).get("name"):
                student_names[user_id] = sub.get("user", {}).get("name")
            submissions[user_id] = body
    
    print(f"   Submissions with content: {len(submissions)}")
    
    if len(submissions) < 2:
        print("❌ Need at least 2 submissions to compare.")
        return
    
    # Run similarity analysis
    print("\n🔍 Analyzing phrase similarities (this may take a moment)...")
    
    matches = check_semantic_similarity_across_submissions(
        submissions, student_names, similarity_threshold=0.6
    )
    
    # Report results
    print("\n" + "="*70)
    print("📊 PHRASE SIMILARITY RESULTS")
    print("="*70)
    print(f"\nAssignment: {assignment_name}")
    print(f"Students analyzed: {len(submissions)}")
    print()
    
    if matches:
        # Sort by total matches
        matches.sort(key=lambda x: x["total_matches"], reverse=True)
        
        print(f"🚨 SIMILAR PHRASE PAIRS FOUND: {len(matches)}")
        print("-"*70)
        
        for match in matches:
            print(f"\n👥 {match['student1']} ↔ {match['student2']}")
            print(f"   Total similarities: {match['total_matches']}")
            
            if match["exact_matches"]:
                print(f"   📌 Exact phrase matches:")
                for phrase in match["exact_matches"][:3]:
                    print(f"      \"{phrase[:60]}...\"" if len(phrase) > 60 else f"      \"{phrase}\"")
            
            if match["reordered_matches"]:
                print(f"   🔄 Same words, different order:")
                for p1, p2 in match["reordered_matches"][:2]:
                    print(f"      \"{p1[:40]}...\" vs \"{p2[:40]}...\"")
            
            if match["similar_phrases"]:
                print(f"   🔗 Similar phrases:")
                for sim in match["similar_phrases"][:2]:
                    print(f"      \"{sim['phrase1'][:40]}...\"")
                    print(f"      \"{sim['phrase2'][:40]}...\" ({int(sim['similarity']*100)}% similar)")
    else:
        print("✅ No significant phrase similarities found across students.")
    
    # Export option
    if matches:
        print()
        export = input("Export results to CSV? (y/n, default=n): ").strip().lower()
        if export == 'y':
            BASE_DIR = get_output_base_dir()
            output_dir = BASE_DIR / "Academic Dishonesty Reports" / "csv"
            output_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_path = output_dir / f"phrase_similarity_{course_id}_{assignment_id}_{timestamp}.csv"
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    "Student 1", "Student 1 ID", "Student 2", "Student 2 ID",
                    "Total Matches", "Exact Matches", "Reordered Matches", "Similar Phrases"
                ])
                for m in matches:
                    writer.writerow([
                        m["student1"], m["student1_id"], m["student2"], m["student2_id"],
                        m["total_matches"],
                        "; ".join(m["exact_matches"][:5]),
                        "; ".join([f"{p1} / {p2}" for p1, p2 in m["reordered_matches"][:3]]),
                        "; ".join([f"{s['phrase1'][:30]}..." for s in m["similar_phrases"][:3]])
                    ])
            
            print(f"✅ Exported to: {csv_path}")


def main():
    """Main entry point for the Academic Dishonesty Flag Generator."""
    print("🎓 Academic Dishonesty Flag Generator (for Canvas)")
    print("Designed for detecting AI-generated content, plagiarism, and bad faith work.")
    print("NO GRADES WILL BE SUBMITTED.\n")
    
    # Main menu - choose between standard analysis or special tools
    print("="*70)
    print("📋 MAIN MENU")
    print("="*70)
    print()
    print("   [1] 🔍 Standard Analysis - Analyze assignments for AI/dishonesty flags")
    print("   [2] 🔑 White Text Detection - Quick check for hidden keyword triggers")
    print("   [3] 📝 Draft Comparison - Compare rough draft vs final draft")
    print("   [4] 🔗 Cross-Student Phrase Analysis - Find similar phrases across students")
    print()
    
    main_choice = input("Select option (1-4, default=1): ").strip() or "1"
    
    try:
        course_id = int(input("\nEnter Course ID: ").strip())
    except ValueError:
        print("❌ Invalid course ID.")
        return

    # Verify API access
    print("🔍 Verifying Canvas API access...")
    test_url = f"{CANVAS_BASE_URL}/api/v1/courses/{course_id}"
    test_resp = requests.get(test_url, headers=HEADERS)
    
    if not test_resp.headers.get('Content-Type', '').startswith('application/json'):
        print("❌ CRITICAL: Received HTML response.")
        print("   Check your token, URL, and internet connection.")
        return
    
    if test_resp.status_code != 200:
        print(f"❌ Could not access course: {test_resp.text}")
        return
    
    course_data = test_resp.json()
    course_name = course_data.get("name", f"Course {course_id}")
    print(f"✅ Connected to: {course_name}\n")
    
    # Route to appropriate analysis mode
    if main_choice == "2":
        run_white_text_mode(course_id, course_name)
        return
    elif main_choice == "3":
        run_draft_comparison_mode(course_id, course_name)
        return
    elif main_choice == "4":
        run_phrase_similarity_mode(course_id, course_name)
        return
    
    # Standard analysis continues below
    # Get students
    students = get_active_students(course_id)
    if not students:
        print("🛑 No active students found. Exiting.")
        return
    
    print(f"✅ Found {len(students)} active students\n")

    # Select assignment type profile for analysis
    global CURRENT_PROFILE
    CURRENT_PROFILE = select_assignment_profile()

    # Ask if user wants to analyze all assignments or specific ones
    print("\n📋 Assignment Selection Options:")
    print("   [1] Analyze ALL assignments in this course")
    print("   [2] Analyze specific assignments by ID")
    print("   [3] Analyze assignments filtered by keyword")
    choice = input("\nChoose an option (1/2/3, default=1): ").strip() or "1"

    assignments = get_all_assignments(course_id)
    if not assignments:
        print("🛑 No assignments found. Exiting.")
        return

    print(f"\n✅ Found {len(assignments)} total assignments in the course")

    if choice == "1":
        # Use all assignments
        print("\n✅ Selected: ALL assignments will be analyzed")
        
    elif choice == "2":
        # Manual assignment ID entry
        assignment_input = input("\nEnter specific Assignment ID(s) (space-separated): ").strip()
        try:
            assignment_ids = [int(x.strip()) for x in assignment_input.split() if x.strip()]
            if not assignment_ids:
                print("❌ No assignment IDs provided. Using all assignments instead.")
            else:
                # Filter assignments list to only include specified IDs
                filtered_assignments = [a for a in assignments if a.get("id") in assignment_ids]
                if len(filtered_assignments) != len(assignment_ids):
                    print(f"⚠️  Some assignment IDs not found. Analyzing {len(filtered_assignments)} valid assignments.")
                assignments = filtered_assignments
        except ValueError:
            print("❌ Invalid assignment ID format. Using all assignments instead.")
            
    elif choice == "3":
        # Filter by keyword
        filter_keyword = input("\nEnter keyword to filter assignments (e.g., 'essay', 'reflection'): ").strip().lower()
        if filter_keyword:
            assignments = [a for a in assignments if filter_keyword in a.get("name", "").lower()]
            print(f"📌 Filtered to {len(assignments)} assignments containing '{filter_keyword}'")
        else:
            print("⚠️  No keyword provided. Using all assignments.")

    # Ask if user wants to analyze discussion forums
    if choice == "1":
        print("\n💬 Fetching ALL discussion topics (auto-selected with 'Analyze ALL')...")
        analyze_discussions = True
        discussion_topics = get_all_discussion_topics(course_id)
        print(f"✅ Found {len(discussion_topics)} discussion topics\n")
    else:
        # Only ask for confirmation and filtering in manual modes ([2] or [3])
        analyze_discussions = input("Also analyze discussion forums for academic dishonesty? (y/n, default=n): ").strip().lower() == 'y'
        if analyze_discussions:
            print("\n💬 Fetching discussion topics...")
            discussion_topics = get_all_discussion_topics(course_id)
            print(f"✅ Found {len(discussion_topics)} discussion topics\n")
            if discussion_topics:
                filter_discussions = input("Filter discussion topics? (y/n, default=n): ").strip().lower() == 'y'
                if filter_discussions:
                    filter_keyword = input("Enter keyword to filter discussions: ").strip().lower()
                    if filter_keyword:
                        discussion_topics = [t for t in discussion_topics if filter_keyword in t.get("title", "").lower()]
                        print(f"📌 Filtered to {len(discussion_topics)} discussions containing '{filter_keyword}'\n")
        else:
            discussion_topics = []

    # Analyze assignments
    print(f"\n🔍 Analyzing assignments ({len(assignments)} total)...")
    all_data = {}
    for idx, assignment in enumerate(assignments, 1):
        assignment_id = assignment.get("id")
        assignment_name = assignment.get("name", f"Assignment {assignment_id}")
        
        print(f"\n📝 [{idx}/{len(assignments)}] Analyzing: {assignment_name}")
        
        student_flags, submission_texts = analyze_assignment(course_id, assignment, students)
        
        # Get student names from submissions
        submissions = get_submissions(course_id, assignment_id)
        student_names = {}
        for user_id, sub in submissions.items():
            user_info = sub.get("user", {})
            student_names[user_id] = user_info.get("name", f"User {user_id}")
        
        all_data[assignment_id] = {
            "name": assignment_name,
            "flags": student_flags,
            "student_names": student_names,
            "submission_texts": submission_texts
        }
    
    print("\n✅ Assignment analysis complete!")
    
    # Analyze discussion topics if requested
    discussion_data = {}
    if analyze_discussions and discussion_topics:
        print("\n🔍 Analyzing discussion topics...")
        
        for topic in discussion_topics:
            topic_id = topic.get("id")
            topic_name = topic.get("title", f"Discussion {topic_id}")
            
            student_flags, post_texts = analyze_discussion_posts(course_id, topic, students)
            
            # Get student names from discussion posts
            student_names = {}
            entries = fetch_discussion_entries(course_id, topic_id)
            for entry in entries:
                user_id = entry.get("user_id")
                if user_id:
                    user_info = entry.get("user", {})
                    student_names[user_id] = user_info.get("name", f"User {user_id}")
            
            discussion_data[topic_id] = {
                "name": topic_name,
                "flags": student_flags,
                "student_names": student_names,
                "post_texts": post_texts
            }
        
        print("\n✅ Discussion forum analysis complete!")

    # Check for cross-assignment similarities
    student_names_global = {}
    for assignment_data in all_data.values():
        student_names_global.update(assignment_data.get("student_names", {}))
    
    # Also include discussion student names
    for topic_data in discussion_data.values():
        student_names_global.update(topic_data.get("student_names", {}))
    
    self_plagiarism, cross_student_plagiarism = check_cross_assignment_similarity(all_data, student_names_global)
    
    # Generate report
    generate_report(course_id, course_name, all_data, self_plagiarism, cross_student_plagiarism, discussion_data)
    
    print("\n" + "="*80)
    print("🏁 Report generation complete!")
    print("="*80)

    # ✅ Final output location instructions (cross-platform)
    # Define output_dir here since generate_report() has its own local scope
    output_dir = get_output_base_dir() / "Academic Dishonesty Reports"
    print(f"\n📁 All reports were exported to: {output_dir}")
    
    system = platform.system()
    if system == "Darwin":
        print(f"\n   🔍 To view in Finder:")
        print(f"       Open Finder → Go → Go to Folder… → paste the path above")
        print(f"\n   💻 To open in Terminal:")
        print(f'       open "{output_dir}"')
    elif system == "Windows":
        print(f"\n   🔍 To view in Explorer:")
        print(f"       Press Win+E → paste the path in the address bar")
        print(f"\n   💻 To open in Command Prompt:")
        print(f'       explorer "{output_dir}"')
    else:
        print(f"\n   🔍 To view in file manager:")
        print(f"       Open your file manager and navigate to the path above")
        print(f"\n   💻 To open in Terminal:")
        print(f'       xdg-open "{output_dir}"')

if __name__ == "__main__":
    main()